"""深市主板固定策略的严格下一交易日回测核心。

本模块复用根目录生产策略中的因子和筛选函数，但为历史回测单独保存每只股票的
长区间完整日线，避免按每个交易日重复联网。
"""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
import hashlib
import json
import math
import os
from pathlib import Path
import time
from typing import Any, Callable, Mapping, Sequence

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import requests

try:  # 支持 `python strategy_backtest/run_backtest.py` 和包导入两种方式。
    from . import factor_batch
    from .runtime_strategy import ROOT_STRATEGY_PATH, strategy_app
except ImportError:  # pragma: no cover - 命令行脚本直接执行时走此分支。
    import factor_batch
    from runtime_strategy import ROOT_STRATEGY_PATH, strategy_app


MODULE_DIR = Path(__file__).resolve().parent
HISTORY_CACHE_ROOT = MODULE_DIR / "data_cache" / "long_history"
FACTOR_CACHE_ROOT = MODULE_DIR / "data_cache" / "strategy_factors"

HISTORY_CACHE_VERSION = 2
LEGACY_HISTORY_CACHE_VERSION = 1
MERGED_HISTORY_CACHE_DIRECTORY = "_merged_bars640_v2"
FACTOR_CACHE_VERSION = 4
# 全量交互回测不能复用固定策略留下的部分因子缓存。此前的试验性快速序列
# 也不满足“每个截至日从固定预热窗口重新起算”的语义，因此单独升级版本。
# v7 保存可配置窗口所需的 MACD 最近金叉日期和年龄。
FULL_FACTOR_CACHE_VERSION = 7
FULL_HISTORY_BAR_LIMIT = 640
FULL_HISTORY_LOOKBACK_DAYS = 420
DEFAULT_CACHE_HOURS = 24.0 * 30.0
DEFAULT_WORKERS = 4
MAX_WORKERS = 6
DEFAULT_FACTOR_WORKERS = max(1, min(8, os.cpu_count() or 1))
MAX_FACTOR_WORKERS = 8
DEFAULT_REQUEST_INTERVAL_SECONDS = 0.25
DEFAULT_TIMEOUT_SECONDS = 15.0

FIXED_TURNOVER_RANGE = (5.0, 10.0)
FIXED_FLOAT_MARKET_CAP_RANGE_YI = (50.0, 200.0)
FIXED_PCT_CHANGE_RANGE = (3.0, 5.0)
FIXED_AMPLITUDE_THRESHOLD = 3.0

RETURN_DETAIL_COLUMNS = (
    "股票代码",
    "交易日期",
    "前一交易日",
    "当日涨跌幅（%）",
)
RETURN_DETAIL_SHEET_NAME = "每日涨跌幅明细"
RETURN_FAILURE_SHEET_NAME = "失败明细"
RETURN_FAILURE_CODE_COLUMN = "股票代码"
HISTORY_ERROR_COLUMNS = (
    "序号",
    "股票代码",
    "股票名称",
    "选股日期",
    "问题类型",
    "失败原因",
)
HISTORY_SCHEMA_COLUMNS = tuple(strategy_app.HISTORY_COLUMNS)
STRATEGY_SNAPSHOT_SIGNATURE = hashlib.sha256(ROOT_STRATEGY_PATH.read_bytes()).hexdigest()
REQUIRED_CACHED_FACTOR_FIELDS = frozenset(
    {
        "数据日期",
        "站上MA5",
        "MA5高于MA20",
        "MACD多头",
        "MACD_DIF",
        "MACD_DEA",
        strategy_app.MACD_GOLDEN_CROSS_DATE_COLUMN,
        strategy_app.MACD_GOLDEN_CROSS_AGE_COLUMN,
        strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_NAME,
        strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_DATE_COLUMN,
        strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_AGE_COLUMN,
        "当日成交额",
        "估算流通市值（亿元）",
        "BIAS20",
        "上影线比例",
        "当日涨跌幅",
        "触及60日高点压力",
        *strategy_app.CANDLESTICK_RISK_FACTOR_COLUMNS.values(),
    }
)


class BacktestDataError(RuntimeError):
    """输入的历史数据或收益数据不满足严格回测要求。"""


class HistoryFetchError(BacktestDataError):
    """完整历史日线获取失败。"""


@dataclass(frozen=True)
class ReturnData:
    """来自每日涨跌幅明细的严格下一交易日收益。"""

    signal_dates: tuple[date, ...]
    next_trade_dates: Mapping[date, date]
    strict_returns: Mapping[tuple[date, str], float]
    failed_return_codes: frozenset[str] = frozenset()


@dataclass(frozen=True)
class HistoryOutcome:
    """单只股票长历史的读取或获取结果。"""

    code: str
    history: pd.DataFrame | None
    source: str | None
    from_cache: bool
    cache_token: str | None
    error: str | None = None


@dataclass(frozen=True)
class _CachedHistory:
    """A validated history cache entry together with its covered window."""

    path: Path
    outcome: HistoryOutcome
    coverage_start: date
    coverage_end: date


HistoryProgressCallback = Callable[[int, int, str, int, int, int], None]


def fixed_strategy_settings() -> tuple[dict[str, bool], dict[str, bool]]:
    """返回本次回测要求的固定选股和风险参数副本。"""

    selected = {key: False for key in strategy_app.SCORING_INDICATOR_KEYS}
    selected.update(
        {
            "above_ma5": True,
            "ma5_above_ma20": True,
            "macd_bullish": True,
            "kdj_healthy_golden_cross_3d": True,
            "amount_at_least_100m": True,
            "float_market_cap_in_range": True,
        }
    )
    selected_risks = {
        "bias_high": True,
        "upper_shadow": True,
        "resistance_60_day": True,
        **{key: True for key in strategy_app.CANDLESTICK_RISK_PATTERN_KEYS},
    }
    return selected, selected_risks


def _as_six_digit_code(value: object) -> str | None:
    """将工作簿中的股票代码规范为六位文本。"""

    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if text.isdigit() and len(text) <= 6:
        return text.zfill(6)
    return None


def _as_date(value: object, *, field_name: str) -> date:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise BacktestDataError(f"{field_name}不是有效日期：{value!r}。")
    return pd.Timestamp(parsed).date()


def load_strict_next_day_returns(workbook_path: Path) -> ReturnData:
    """读取收益工作簿，并只保留严格相邻市场交易日的单日收益。

    例如某股票停牌后在更晚日期复牌，虽然该行可能有累计涨跌幅，但前一交易日
    不等于选股日，就绝不作为“第二天”的真实收益。
    """

    if not workbook_path.is_file():
        raise FileNotFoundError(f"未找到每日涨跌幅文件：{workbook_path}。")
    try:
        with pd.ExcelFile(workbook_path) as workbook:
            if RETURN_DETAIL_SHEET_NAME not in workbook.sheet_names:
                raise BacktestDataError("收益工作簿中没有“每日涨跌幅明细”工作表。")
            detail = pd.read_excel(
                workbook,
                sheet_name=RETURN_DETAIL_SHEET_NAME,
                dtype={"股票代码": str},
            )
            failed_return_codes = frozenset()
            if RETURN_FAILURE_SHEET_NAME in workbook.sheet_names:
                failures = pd.read_excel(
                    workbook,
                    sheet_name=RETURN_FAILURE_SHEET_NAME,
                    dtype={RETURN_FAILURE_CODE_COLUMN: str},
                )
                if RETURN_FAILURE_CODE_COLUMN not in failures.columns:
                    raise BacktestDataError("“失败明细”工作表缺少“股票代码”列。")
                failed_return_codes = frozenset(
                    code
                    for value in failures[RETURN_FAILURE_CODE_COLUMN]
                    if (code := _as_six_digit_code(value)) is not None
                )
    except ValueError as exc:
        raise BacktestDataError("无法读取收益工作簿。") from exc

    missing = set(RETURN_DETAIL_COLUMNS).difference(detail.columns)
    if missing:
        missing_text = "、".join(sorted(missing))
        raise BacktestDataError(f"收益工作簿缺少必要列：{missing_text}。")

    returns = detail.loc[:, RETURN_DETAIL_COLUMNS].copy()
    returns["股票代码"] = returns["股票代码"].map(_as_six_digit_code)
    returns["交易日期"] = pd.to_datetime(returns["交易日期"], errors="coerce").dt.normalize()
    returns["前一交易日"] = pd.to_datetime(returns["前一交易日"], errors="coerce").dt.normalize()
    returns["当日涨跌幅（%）"] = pd.to_numeric(returns["当日涨跌幅（%）"], errors="coerce")
    returns = returns.dropna(
        subset=["股票代码", "交易日期", "前一交易日", "当日涨跌幅（%）"]
    )
    returns = returns.drop_duplicates(subset=["股票代码", "交易日期"], keep="last")
    if returns.empty:
        raise BacktestDataError("收益工作簿没有可用的每日涨跌幅记录。")

    market_dates = tuple(sorted(pd.Timestamp(value).date() for value in returns["交易日期"].unique()))
    if len(market_dates) < 2:
        raise BacktestDataError("收益工作簿至少需要两个实际交易日。")
    signal_dates = market_dates[:-1]
    next_trade_dates = dict(zip(signal_dates, market_dates[1:], strict=True))

    expected_next = returns["前一交易日"].dt.date.map(next_trade_dates)
    strict = returns.loc[
        expected_next.notna() & returns["交易日期"].dt.date.eq(expected_next)
    ].copy()
    strict_returns = {
        (pd.Timestamp(row["前一交易日"]).date(), str(row["股票代码"])): float(
            row["当日涨跌幅（%）"]
        )
        for _, row in strict.iterrows()
    }
    return ReturnData(
        signal_dates=signal_dates,
        next_trade_dates=next_trade_dates,
        strict_returns=strict_returns,
        failed_return_codes=failed_return_codes,
    )


def build_next_day_open_to_close_returns(
    return_data: ReturnData,
    history_outcomes: Mapping[str, HistoryOutcome],
) -> dict[tuple[date, str], float]:
    """从完整日线生成下一交易日开盘至收盘的可执行收益。"""

    needed_days = set(return_data.next_trade_dates.values())
    realized_returns: dict[tuple[date, str], float] = {}
    for code, outcome in history_outcomes.items():
        if outcome.history is None or outcome.history.empty:
            continue
        bars_by_day: dict[date, tuple[float, float]] = {}
        opens = pd.to_numeric(outcome.history["open"], errors="coerce")
        closes = pd.to_numeric(outcome.history["close"], errors="coerce")
        for raw_day, open_price, close_price in zip(
            outcome.history["date"], opens, closes, strict=True
        ):
            trading_day = pd.Timestamp(raw_day).date()
            if trading_day not in needed_days:
                continue
            if (
                pd.isna(open_price)
                or pd.isna(close_price)
                or float(open_price) <= 0.0
            ):
                continue
            bars_by_day[trading_day] = (float(open_price), float(close_price))
        for signal_day, next_day in return_data.next_trade_dates.items():
            prices = bars_by_day.get(next_day)
            if prices is None:
                continue
            open_price, close_price = prices
            realized_returns[(signal_day, str(code))] = (
                close_price / open_price - 1.0
            ) * 100.0
    return realized_returns


def history_cache_key(first_signal_date: date, last_market_date: date) -> str:
    """Return the factor-cache key for a requested backtest window."""

    return (
        f"{first_signal_date.isoformat()}_{last_market_date.isoformat()}"
        f"_bars{FULL_HISTORY_BAR_LIMIT}"
    )


def _history_cache_path(cache_key: str, code: str) -> Path:
    """Return the shared, merged long-history cache path for one stock."""

    del cache_key
    return HISTORY_CACHE_ROOT / MERGED_HISTORY_CACHE_DIRECTORY / f"{code}.json"


def _legacy_history_cache_path(cache_key: str, code: str) -> Path:
    """Return a pre-v2 range-specific cache path for migration reads only."""

    return HISTORY_CACHE_ROOT / cache_key / f"{code}.json"


def _factor_cache_path(cache_key: str, code: str) -> Path:
    return FACTOR_CACHE_ROOT / cache_key / f"{code}.json"


def _full_factor_cache_key(cache_key: str) -> str:
    """返回全量交互回测专用的因子缓存键。

    固定策略缓存只覆盖其预筛选通过的股票日，不能作为任意勾选项回测的全量
    面板。单独的键还能使历史上错误的连续序列结果无法被本模块读取。
    """

    return f"{cache_key}_all_factors_v{FULL_FACTOR_CACHE_VERSION}"


def _cache_is_fresh(path: Path, cache_hours: float) -> bool:
    if cache_hours <= 0 or not path.is_file():
        return False
    age_seconds = max(0.0, time.time() - path.stat().st_mtime)
    return age_seconds <= float(cache_hours) * 3600.0


def _history_cache_token(payload: Mapping[str, object], path: Path) -> str:
    saved_at = str(payload.get("saved_at") or "")
    return f"{saved_at}:{path.stat().st_mtime_ns}"


def _as_cache_date(value: object) -> date | None:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed).date()


def _history_cache_key_bounds(cache_key: str) -> tuple[date, date] | None:
    """Parse the date bounds encoded by the former range-specific cache key."""

    parts = cache_key.split("_")
    if len(parts) != 3 or not parts[2].startswith("bars"):
        return None
    first_signal_date = _as_cache_date(parts[0])
    end_date = _as_cache_date(parts[1])
    if (
        first_signal_date is None
        or end_date is None
        or first_signal_date > end_date
    ):
        return None
    return first_signal_date, end_date


def _legacy_history_cache_keys(
    first_signal_date: date,
    end_date: date,
) -> tuple[str, ...]:
    """Order legacy cache directories by their likely coverage of this request."""

    try:
        directory_names = [
            path.name
            for path in HISTORY_CACHE_ROOT.iterdir()
            if path.is_dir() and path.name != MERGED_HISTORY_CACHE_DIRECTORY
        ]
    except OSError:
        return ()

    needed_start = first_signal_date - timedelta(days=FULL_HISTORY_LOOKBACK_DAYS)

    def sort_key(cache_key: str) -> tuple[int, int, int, int, str]:
        bounds = _history_cache_key_bounds(cache_key)
        if bounds is None:
            return (1, 1, 0, 0, cache_key)
        cached_first_signal_date, cached_end_date = bounds
        cached_start = cached_first_signal_date - timedelta(
            days=FULL_HISTORY_LOOKBACK_DAYS
        )
        return (
            0 if cached_start <= needed_start and cached_end_date >= end_date else 1,
            0 if cached_end_date >= end_date else 1,
            -min(cached_end_date, end_date).toordinal(),
            cached_start.toordinal(),
            cache_key,
        )

    return tuple(sorted(directory_names, key=sort_key))


def _read_history_cache_entry(
    code: str,
    *,
    path: Path,
    cache_hours: float,
    allow_stale: bool = False,
) -> _CachedHistory | None:
    """Read and validate one shared or legacy long-history cache file."""

    if allow_stale:
        if not path.is_file():
            return None
    elif not _cache_is_fresh(path, cache_hours):
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, Mapping):
            return None
        if payload.get("version") not in (
            LEGACY_HISTORY_CACHE_VERSION,
            HISTORY_CACHE_VERSION,
        ):
            return None
        if payload.get("code") != code:
            return None
        coverage_end = _as_cache_date(payload.get("end_date"))
        if coverage_end is None:
            return None
        raw_bars = payload.get("bars")
        if not isinstance(raw_bars, list):
            return None
        cached_columns = payload.get("history_columns")
        if cached_columns is None:
            # 旧缓存没有写入字段元数据时，只有每一根日线仍具备完整字段才兼容。
            if any(
                not isinstance(bar, Mapping)
                or not set(HISTORY_SCHEMA_COLUMNS).issubset(bar)
                for bar in raw_bars
            ):
                return None
        elif tuple(cached_columns) != HISTORY_SCHEMA_COLUMNS:
            return None
        history = strategy_app._normalize_history_frame(pd.DataFrame(raw_bars))
        history = history.loc[
            history["date"].le(pd.Timestamp(coverage_end))
        ].reset_index(drop=True)
        if history.empty:
            return None
        actual_coverage_start = pd.Timestamp(history["date"].iloc[0]).date()
        coverage_start = _as_cache_date(payload.get("coverage_start"))
        if coverage_start is None:
            cached_first_signal_date = _as_cache_date(
                payload.get("first_signal_date")
            )
            coverage_start = (
                cached_first_signal_date
                - timedelta(days=FULL_HISTORY_LOOKBACK_DAYS)
                if cached_first_signal_date is not None
                else actual_coverage_start
            )
        # A provider can return more history than the requested warmup window.
        # Reuse those bars instead of downloading the already cached prefix again.
        coverage_start = min(coverage_start, actual_coverage_start)
        if coverage_start > coverage_end:
            return None
        return _CachedHistory(
            path=path,
            outcome=HistoryOutcome(
                code=code,
                history=history,
                source=str(payload.get("source") or "本地长历史缓存"),
                from_cache=True,
                cache_token=_history_cache_token(payload, path),
            ),
            coverage_start=coverage_start,
            coverage_end=coverage_end,
        )
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return None


def _cached_history_covers(
    cached: _CachedHistory,
    *,
    history_start: date,
    end_date: date,
) -> bool:
    return cached.coverage_start <= history_start and cached.coverage_end >= end_date


def _cached_history_score(
    cached: _CachedHistory,
    *,
    history_start: date,
    end_date: date,
) -> tuple[int, int, int, int]:
    overlap_start = max(cached.coverage_start, history_start)
    overlap_end = min(cached.coverage_end, end_date)
    overlap_days = max(0, (overlap_end - overlap_start).days + 1)
    return (
        int(_cached_history_covers(cached, history_start=history_start, end_date=end_date)),
        overlap_days,
        cached.coverage_end.toordinal(),
        -cached.coverage_start.toordinal(),
    )


def _find_cached_history(
    code: str,
    *,
    cache_key: str,
    first_signal_date: date,
    end_date: date,
    cache_hours: float,
    legacy_cache_keys: Sequence[str],
    allow_stale: bool = False,
) -> _CachedHistory | None:
    """Use the merged cache first, then migrate the best usable legacy entry."""

    history_start = first_signal_date - timedelta(days=FULL_HISTORY_LOOKBACK_DAYS)
    cached = _read_history_cache_entry(
        code,
        path=_history_cache_path(cache_key, code),
        cache_hours=cache_hours,
        allow_stale=allow_stale,
    )
    if cached is not None and _cached_history_covers(
        cached,
        history_start=history_start,
        end_date=end_date,
    ):
        return cached

    best_cached = cached
    for legacy_cache_key in legacy_cache_keys:
        legacy_cached = _read_history_cache_entry(
            code,
            path=_legacy_history_cache_path(legacy_cache_key, code),
            cache_hours=cache_hours,
            allow_stale=allow_stale,
        )
        if legacy_cached is None:
            continue
        if _cached_history_covers(
            legacy_cached,
            history_start=history_start,
            end_date=end_date,
        ):
            return legacy_cached
        if best_cached is None or _cached_history_score(
            legacy_cached,
            history_start=history_start,
            end_date=end_date,
        ) > _cached_history_score(
            best_cached,
            history_start=history_start,
            end_date=end_date,
        ):
            best_cached = legacy_cached
    return best_cached


def _read_history_cache(
    code: str,
    *,
    cache_key: str,
    end_date: date,
    cache_hours: float,
    first_signal_date: date | None = None,
) -> HistoryOutcome | None:
    """Read a merged cache entry when it covers the requested history window."""

    cached = _read_history_cache_entry(
        code,
        path=_history_cache_path(cache_key, code),
        cache_hours=cache_hours,
    )
    if cached is None:
        cached = _read_history_cache_entry(
            code,
            path=_legacy_history_cache_path(cache_key, code),
            cache_hours=cache_hours,
        )
    if cached is None or cached.coverage_end < end_date:
        return None
    if first_signal_date is not None and cached.coverage_start > (
        first_signal_date - timedelta(days=FULL_HISTORY_LOOKBACK_DAYS)
    ):
        return None
    history = cached.outcome.history
    if history is None:
        return None
    history = history.loc[history["date"].le(pd.Timestamp(end_date))].reset_index(
        drop=True
    )
    if history.empty:
        return None
    return HistoryOutcome(
        code=code,
        history=history,
        source=cached.outcome.source,
        from_cache=True,
        cache_token=cached.outcome.cache_token,
    )


def _merge_history_frames(*frames: pd.DataFrame) -> pd.DataFrame:
    usable_frames = [frame for frame in frames if not frame.empty]
    if not usable_frames:
        return pd.DataFrame(columns=HISTORY_SCHEMA_COLUMNS)
    return strategy_app._normalize_history_frame(
        pd.concat(usable_frames, ignore_index=True)
    )


def _write_history_cache(
    code: str,
    history: pd.DataFrame,
    source: str,
    *,
    cache_key: str,
    first_signal_date: date,
    end_date: date,
    coverage_start: date | None = None,
) -> str:
    """Atomically persist the merged history cache and return its token."""

    target = _history_cache_path(cache_key, code)
    target.parent.mkdir(parents=True, exist_ok=True)
    saved_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    coverage_start = coverage_start or (
        first_signal_date - timedelta(days=FULL_HISTORY_LOOKBACK_DAYS)
    )
    payload: dict[str, object] = {
        "version": HISTORY_CACHE_VERSION,
        "code": code,
        "first_signal_date": first_signal_date.isoformat(),
        "coverage_start": coverage_start.isoformat(),
        "end_date": end_date.isoformat(),
        "saved_at": saved_at,
        "source": source,
        "history_columns": list(HISTORY_SCHEMA_COLUMNS),
        "bars": strategy_app._history_to_records(history),
    }
    temporary = target.with_name(
        f"{target.name}.{os.getpid()}.{time.time_ns()}.tmp"
    )
    try:
        temporary.write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        temporary.replace(target)
    finally:
        try:
            temporary.unlink(missing_ok=True)
        except OSError:
            pass
    return _history_cache_token(payload, target)


def _parse_tencent_rich_full_history(raw_text: str, symbol: str, end_date: date) -> pd.DataFrame:
    """解析腾讯增强前复权日 K，保留完整历史而非原应用的短窗口。"""

    json_text = raw_text.strip()
    if "=" in json_text:
        json_text = json_text.split("=", 1)[1].strip()
    json_text = json_text.rstrip(";").strip()
    try:
        payload = json.loads(json_text)
    except json.JSONDecodeError as exc:
        raise HistoryFetchError("腾讯增强日 K 响应不是有效 JSON。") from exc
    if not isinstance(payload, Mapping) or payload.get("code") not in (None, 0, "0"):
        raise HistoryFetchError("腾讯增强日 K 响应状态异常。")
    data = payload.get("data")
    if not isinstance(data, Mapping):
        raise HistoryFetchError("腾讯增强日 K 响应没有 data 对象。")
    security_data = data.get(symbol)
    if not isinstance(security_data, Mapping):
        raise HistoryFetchError(f"腾讯增强日 K 响应没有 {symbol} 数据。")
    raw_klines = security_data.get("qfqday")
    if not isinstance(raw_klines, list):
        raw_klines = security_data.get("day")
    if not isinstance(raw_klines, list):
        raise HistoryFetchError("腾讯增强日 K 响应缺少前复权日线。")

    rows: list[dict[str, object]] = []
    for raw_kline in raw_klines:
        if not isinstance(raw_kline, (list, tuple)) or len(raw_kline) < 9:
            continue
        amount_in_ten_thousand = strategy_app._coerce_number(raw_kline[8])
        rows.append(
            {
                "date": raw_kline[0],
                "open": strategy_app._coerce_number(raw_kline[1]),
                "close": strategy_app._coerce_number(raw_kline[2]),
                "high": strategy_app._coerce_number(raw_kline[3]),
                "low": strategy_app._coerce_number(raw_kline[4]),
                "volume": strategy_app._coerce_number(raw_kline[5]),
                "amount": (
                    amount_in_ten_thousand * 10_000
                    if amount_in_ten_thousand is not None
                    else None
                ),
                "amplitude": None,
                "pct_change": None,
                "turnover": strategy_app._coerce_number(raw_kline[7]),
            }
        )
    history = strategy_app._normalize_history_frame(pd.DataFrame(rows))
    history = history.loc[history["date"].le(pd.Timestamp(end_date))].reset_index(drop=True)
    if history.empty:
        raise HistoryFetchError("腾讯增强日 K 没有有效日线。")
    return history


def _parse_eastmoney_full_history(payload: Mapping[str, Any], end_date: date) -> pd.DataFrame:
    """解析东方财富前复权日 K，保留完整历史。"""

    if str(payload.get("rc", "0")) not in {"0", "None"} and payload.get("rc") not in (0, None):
        raise HistoryFetchError(f"东方财富接口返回 rc={payload.get('rc')!r}。")
    data = payload.get("data")
    if not isinstance(data, Mapping):
        raise HistoryFetchError("东方财富响应没有 data 对象。")
    raw_klines = data.get("klines")
    if not isinstance(raw_klines, list):
        raise HistoryFetchError("东方财富响应没有日 K 列表。")

    rows: list[dict[str, object]] = []
    for raw_kline in raw_klines:
        if not isinstance(raw_kline, str):
            continue
        fields = raw_kline.split(",")
        if len(fields) < 11:
            continue
        rows.append(
            {
                "date": fields[0],
                "open": strategy_app._coerce_number(fields[1]),
                "close": strategy_app._coerce_number(fields[2]),
                "high": strategy_app._coerce_number(fields[3]),
                "low": strategy_app._coerce_number(fields[4]),
                "volume": strategy_app._coerce_number(fields[5]),
                "amount": strategy_app._coerce_number(fields[6]),
                "amplitude": strategy_app._coerce_number(fields[7]),
                "pct_change": strategy_app._coerce_number(fields[8]),
                "turnover": strategy_app._coerce_number(fields[10]),
            }
        )
    history = strategy_app._normalize_history_frame(pd.DataFrame(rows))
    history = history.loc[history["date"].le(pd.Timestamp(end_date))].reset_index(drop=True)
    if history.empty:
        raise HistoryFetchError("东方财富没有有效日线。")
    return history


def _fetch_tencent_full_history(
    code: str,
    *,
    history_start: date,
    end_date: date,
    limiter: strategy_app.RequestRateLimiter,
    timeout_seconds: float,
) -> pd.DataFrame:
    """请求腾讯增强日 K；640 根覆盖本次回测所需的预热和测试期。"""

    symbol = f"sz{code}"
    parameters = {
        "_var": "kline_dayqfq",
        "param": (
            f"{symbol},day,{history_start.isoformat()},{end_date.isoformat()},"
            f"{FULL_HISTORY_BAR_LIMIT},qfq"
        ),
        "r": f"{time.time():.6f}",
    }
    problems: list[str] = []
    for attempt in range(1, strategy_app.MAX_RETRIES + 1):
        for endpoint in strategy_app.TENCENT_RICH_DAILY_KLINE_URLS:
            try:
                limiter.wait()
                response = strategy_app._request_session().get(
                    endpoint,
                    params=parameters,
                    timeout=timeout_seconds,
                )
                response.raise_for_status()
                return _parse_tencent_rich_full_history(response.text, symbol, end_date)
            except requests.RequestException as exc:
                if strategy_app._is_rate_limited(exc):
                    limiter.penalize(1.5 * attempt)
                problems.append(f"第 {attempt} 次 {endpoint}：{exc}")
            except (HistoryFetchError, TypeError, ValueError) as exc:
                problems.append(f"第 {attempt} 次 {endpoint}：{exc}")
        if attempt < strategy_app.MAX_RETRIES:
            time.sleep(0.8 * attempt)
    raise HistoryFetchError("腾讯增强日 K 失败：" + "；".join(problems))


def _fetch_eastmoney_full_history(
    code: str,
    *,
    history_start: date,
    end_date: date,
    limiter: strategy_app.RequestRateLimiter,
    timeout_seconds: float,
) -> pd.DataFrame:
    """请求东方财富完整前复权日 K。"""

    parameters = {
        "secid": f"0.{code}",
        "klt": "101",
        "fqt": "1",
        "beg": history_start.strftime("%Y%m%d"),
        "end": end_date.strftime("%Y%m%d"),
        "lmt": str(FULL_HISTORY_BAR_LIMIT),
        "fields1": strategy_app.EASTMONEY_FIELDS1,
        "fields2": strategy_app.EASTMONEY_FIELDS2,
    }
    problems: list[str] = []
    for attempt in range(1, strategy_app.MAX_RETRIES + 1):
        try:
            limiter.wait()
            response = strategy_app._request_session().get(
                strategy_app.EASTMONEY_DAILY_KLINE_URL,
                params=parameters,
                timeout=timeout_seconds,
            )
            response.raise_for_status()
            return _parse_eastmoney_full_history(response.json(), end_date)
        except requests.RequestException as exc:
            if strategy_app._is_rate_limited(exc):
                limiter.penalize(1.5 * attempt)
            problems.append(f"第 {attempt} 次：{exc}")
        except (HistoryFetchError, TypeError, ValueError) as exc:
            problems.append(f"第 {attempt} 次：{exc}")
        if attempt < strategy_app.MAX_RETRIES:
            time.sleep(0.8 * attempt)
    raise HistoryFetchError("东方财富日 K 失败：" + "；".join(problems))


def _fetch_history_range(
    code: str,
    *,
    history_start: date,
    end_date: date,
    limiter: strategy_app.RequestRateLimiter,
    timeout_seconds: float,
) -> tuple[pd.DataFrame, str]:
    """Fetch one missing history interval, preferring Eastmoney as before."""

    try:
        return (
            _fetch_eastmoney_full_history(
                code,
                history_start=history_start,
                end_date=end_date,
                limiter=limiter,
                timeout_seconds=timeout_seconds,
            ),
            "东方财富公开日 K（前复权）",
        )
    except HistoryFetchError as eastmoney_error:
        try:
            return (
                _fetch_tencent_full_history(
                    code,
                    history_start=history_start,
                    end_date=end_date,
                    limiter=limiter,
                    timeout_seconds=timeout_seconds,
                ),
                "腾讯增强前复权日 K 回退",
            )
        except HistoryFetchError as tencent_error:
            raise HistoryFetchError(
                f"东方财富失败：{eastmoney_error}；腾讯回退失败：{tencent_error}"
            ) from tencent_error


def _history_outcome_from_cached(
    cached: _CachedHistory,
    *,
    end_date: date,
    cache_token: str | None = None,
) -> HistoryOutcome | None:
    """Limit a cached history to the requested as-of date for factor calculation."""

    history = cached.outcome.history
    if history is None:
        return None
    history = history.loc[history["date"].le(pd.Timestamp(end_date))].reset_index(
        drop=True
    )
    if history.empty:
        return None
    return HistoryOutcome(
        code=cached.outcome.code,
        history=history,
        source=cached.outcome.source,
        from_cache=True,
        cache_token=cache_token or cached.outcome.cache_token,
    )


def _load_one_history(
    code: str,
    *,
    cache_key: str,
    first_signal_date: date,
    end_date: date,
    cache_hours: float,
    force_refresh: bool,
    limiter: strategy_app.RequestRateLimiter,
    timeout_seconds: float,
    legacy_cache_keys: Sequence[str] | None = None,
) -> HistoryOutcome:
    """Load merged history and request only the selected window's missing ranges."""

    history_start = first_signal_date - timedelta(days=FULL_HISTORY_LOOKBACK_DAYS)
    canonical_path = _history_cache_path(cache_key, code)
    cached: _CachedHistory | None = None
    if cache_hours > 0:
        cached = _find_cached_history(
            code,
            cache_key=cache_key,
            first_signal_date=first_signal_date,
            end_date=end_date,
            cache_hours=cache_hours,
            legacy_cache_keys=(
                ()
                if force_refresh
                else (
                    tuple(legacy_cache_keys)
                    if legacy_cache_keys is not None
                    else _legacy_history_cache_keys(first_signal_date, end_date)
                )
            ),
            allow_stale=force_refresh,
        )

    if (
        not force_refresh
        and cached is not None
        and _cached_history_covers(
            cached,
            history_start=history_start,
            end_date=end_date,
        )
    ):
        if cached.path == canonical_path:
            outcome = _history_outcome_from_cached(cached, end_date=end_date)
            if outcome is not None:
                return outcome
        else:
            cache_token = _write_history_cache(
                code,
                cached.outcome.history,
                cached.outcome.source or "本地长历史缓存",
                cache_key=cache_key,
                first_signal_date=first_signal_date,
                end_date=cached.coverage_end,
                coverage_start=cached.coverage_start,
            )
            outcome = _history_outcome_from_cached(
                cached,
                end_date=end_date,
                cache_token=cache_token,
            )
            if outcome is not None:
                return outcome

    missing_ranges: list[tuple[date, date]]
    if force_refresh or cached is None:
        missing_ranges = [(history_start, end_date)]
    else:
        missing_ranges = []
        if history_start < cached.coverage_start:
            missing_ranges.append(
                (history_start, cached.coverage_start - timedelta(days=1))
            )
        if cached.coverage_end < end_date:
            missing_ranges.append(
                (cached.coverage_end + timedelta(days=1), end_date)
            )

    fetched_histories: list[pd.DataFrame] = []
    fetched_sources: list[str] = []
    try:
        for missing_start, missing_end in missing_ranges:
            if missing_start > missing_end:
                continue
            history, source = _fetch_history_range(
                code,
                history_start=missing_start,
                end_date=missing_end,
                limiter=limiter,
                timeout_seconds=timeout_seconds,
            )
            fetched_histories.append(history)
            fetched_sources.append(source)
    except HistoryFetchError as error:
        return HistoryOutcome(
            code=code,
            history=None,
            source=None,
            from_cache=False,
            cache_token=None,
            error=str(error),
        )

    if cached is None:
        merged_history = _merge_history_frames(*fetched_histories)
        coverage_start = history_start
        coverage_end = end_date
        source = fetched_sources[-1]
    else:
        try:
            merged_history = _merge_history_frames(
                cached.outcome.history, *fetched_histories
            )
        except (TypeError, ValueError) as error:
            return HistoryOutcome(
                code=code,
                history=None,
                source=None,
                from_cache=False,
                cache_token=None,
                error=f"合并本地长历史缓存失败：{error}",
            )
        coverage_start = min(cached.coverage_start, history_start)
        coverage_end = max(cached.coverage_end, end_date)
        source = "; ".join(
            source_part
            for source_part in (cached.outcome.source, *fetched_sources)
            if source_part
        )

    if merged_history.empty:
        return HistoryOutcome(
            code=code,
            history=None,
            source=None,
            from_cache=False,
            cache_token=None,
            error="未获取到可用的长历史日线。",
        )

    cache_token = None
    if cache_hours > 0:
        cache_token = _write_history_cache(
            code,
            merged_history,
            source,
            cache_key=cache_key,
            first_signal_date=first_signal_date,
            end_date=coverage_end,
            coverage_start=coverage_start,
        )
    history = merged_history.loc[
        merged_history["date"].le(pd.Timestamp(end_date))
    ].reset_index(drop=True)
    return HistoryOutcome(
        code=code,
        history=history,
        source=source,
        from_cache=False,
        cache_token=cache_token,
    )


def collect_full_histories(
    companies: pd.DataFrame,
    *,
    first_signal_date: date,
    end_date: date,
    cache_hours: float,
    force_refresh: bool,
    workers: int,
    request_interval_seconds: float,
    timeout_seconds: float,
    progress_callback: HistoryProgressCallback | None = None,
) -> tuple[dict[str, HistoryOutcome], pd.DataFrame, dict[str, int], str]:
    """并发加载所有股票的长历史，并只补齐缓存覆盖范围外的日期。"""

    if not 1 <= int(workers) <= MAX_WORKERS:
        raise ValueError(f"并发数必须在 1 至 {MAX_WORKERS} 之间。")
    if cache_hours < 0:
        raise ValueError("缓存有效小时数不能为负数。")
    if request_interval_seconds < 0:
        raise ValueError("请求最小间隔不能为负数。")
    if timeout_seconds <= 0:
        raise ValueError("请求超时必须为正数。")
    if companies.empty:
        raise BacktestDataError("没有可处理的股票。")

    cache_key = history_cache_key(first_signal_date, end_date)
    legacy_cache_keys = (
        ()
        if force_refresh or cache_hours <= 0
        else _legacy_history_cache_keys(first_signal_date, end_date)
    )
    records = companies.to_dict("records")
    limiter = strategy_app.RequestRateLimiter(float(request_interval_seconds))
    outcomes: dict[str, HistoryOutcome] = {}
    error_rows: list[dict[str, object]] = []
    completed = cache_hits = succeeded = failed = warmup_insufficient = 0

    with ThreadPoolExecutor(max_workers=int(workers)) as executor:
        futures = {
            executor.submit(
                _load_one_history,
                str(record["股票代码"]),
                cache_key=cache_key,
                first_signal_date=first_signal_date,
                end_date=end_date,
                cache_hours=float(cache_hours),
                force_refresh=force_refresh,
                limiter=limiter,
                timeout_seconds=float(timeout_seconds),
                legacy_cache_keys=legacy_cache_keys,
            ): record
            for record in records
        }
        for future in as_completed(futures):
            record = futures[future]
            code = str(record["股票代码"])
            completed += 1
            try:
                outcome = future.result()
            except Exception as exc:  # 防御性兜底，单只失败不应中断整批回测。
                outcome = HistoryOutcome(
                    code=code,
                    history=None,
                    source=None,
                    from_cache=False,
                    cache_token=None,
                    error=f"工作线程异常：{exc}",
                )
            outcomes[code] = outcome
            if outcome.from_cache:
                cache_hits += 1
            if outcome.history is None:
                failed += 1
                error_rows.append(
                    {
                        "序号": record["序号"],
                        "股票代码": code,
                        "股票名称": record["股票名称"],
                        "选股日期": None,
                        "问题类型": "长历史获取失败",
                        "失败原因": outcome.error or "未返回历史行情。",
                    }
                )
            else:
                succeeded += 1
                if len(outcome.history) < strategy_app.MIN_REQUIRED_BARS:
                    warmup_insufficient += 1
                    short_history_error = (
                        f"截至回测末日仅有 {len(outcome.history)} 根有效日线，"
                        f"少于 KDJ 所需的 {strategy_app.MIN_REQUIRED_BARS} 根。"
                    )
                    error_rows.append(
                        {
                            "序号": record["序号"],
                            "股票代码": code,
                            "股票名称": record["股票名称"],
                            "选股日期": None,
                            "问题类型": "历史预热不足",
                            "失败原因": short_history_error,
                        }
                    )
                    # 预热不足的数据保留在问题明细中，但不能再进入任何因子计算器；
                    # 否则会在每个选股日重复生成 KDJ 120 天不足错误。
                    outcomes[code] = HistoryOutcome(
                        code=code,
                        history=None,
                        source=outcome.source,
                        from_cache=outcome.from_cache,
                        cache_token=None,
                        error=short_history_error,
                    )
            if progress_callback is not None:
                progress_callback(completed, len(records), code, cache_hits, succeeded, failed)

    errors = pd.DataFrame(error_rows, columns=HISTORY_ERROR_COLUMNS)
    if not errors.empty:
        errors = errors.sort_values("序号", kind="stable").reset_index(drop=True)
    return outcomes, errors, {
        "股票总数": len(records),
        "历史成功": succeeded,
        "历史失败": failed,
        "历史获取失败股票数": failed,
        "历史预热不足股票数": warmup_insufficient,
        "历史自动剔除股票数": failed + warmup_insufficient,
        "历史有效股票数": succeeded - warmup_insufficient,
        "历史缓存命中": cache_hits,
    }, cache_key


def _basic_prefilter_dates(history: pd.DataFrame, signal_dates: Sequence[date]) -> set[date]:
    """快速判断完全策略中不依赖 KDJ/MACD 的必要条件。

    这些条件均是最终“全部满足”规则的必要条件，因此预筛失败的股票不会影响
    选股结果。完整因子仍由原应用逐日计算，确保 KDJ 和风险规则口径不变。
    """

    bars = history.copy()
    bars["ma5"] = bars["close"].rolling(5, min_periods=5).mean()
    bars["ma20"] = bars["close"].rolling(20, min_periods=20).mean()
    amount = pd.to_numeric(bars["amount"], errors="coerce")
    turnover = pd.to_numeric(bars["turnover"], errors="coerce")
    market_cap_yi = amount * 100.0 / turnover / 100_000_000.0
    signal_timestamps = {pd.Timestamp(item).normalize() for item in signal_dates}
    mask = (
        bars["date"].dt.normalize().isin(signal_timestamps)
        & bars["close"].gt(bars["ma5"])
        & bars["ma5"].gt(bars["ma20"])
        & amount.ge(100_000_000.0)
        & turnover.gt(0.0)
        & market_cap_yi.between(
            FIXED_FLOAT_MARKET_CAP_RANGE_YI[0],
            FIXED_FLOAT_MARKET_CAP_RANGE_YI[1],
            inclusive="both",
        )
    )
    return {pd.Timestamp(value).date() for value in bars.loc[mask, "date"]}


def _json_safe(value: object) -> object:
    """将 pandas 标量和缺失值转换成 JSON 可持久化的数据。"""

    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (str, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value).isoformat()
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        return _json_safe(value.item())
    return str(value)


def _read_factor_cache(
    code: str,
    *,
    cache_key: str,
    history_token: str | None,
    cache_hours: float,
    cache_version: int = FACTOR_CACHE_VERSION,
) -> dict[str, dict[str, object]]:
    """读取与当前长历史完全对应的精算因子缓存。"""

    if not history_token:
        return {}
    target = _factor_cache_path(cache_key, code)
    if not _cache_is_fresh(target, cache_hours):
        return {}
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
        if not isinstance(payload, Mapping):
            return {}
        if payload.get("version") != cache_version:
            return {}
        if payload.get("history_token") != history_token:
            return {}
        if payload.get("strategy_snapshot_signature") != STRATEGY_SNAPSHOT_SIGNATURE:
            return {}
        raw_factors = payload.get("factors")
        if not isinstance(raw_factors, Mapping):
            return {}
        return {
            str(day): dict(values)
            for day, values in raw_factors.items()
            if (
                isinstance(day, str)
                and isinstance(values, Mapping)
                and REQUIRED_CACHED_FACTOR_FIELDS.issubset(values)
            )
        }
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return {}


def _write_factor_cache(
    code: str,
    factors: Mapping[str, Mapping[str, object]],
    *,
    cache_key: str,
    history_token: str | None,
    cache_version: int = FACTOR_CACHE_VERSION,
) -> None:
    """写入单只股票在本回测区间内已精算的因子。"""

    if not history_token:
        return
    target = _factor_cache_path(cache_key, code)
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": cache_version,
        "history_token": history_token,
        "strategy_snapshot_signature": STRATEGY_SNAPSHOT_SIGNATURE,
        "saved_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "factors": {
            str(day): {str(key): _json_safe(value) for key, value in values.items()}
            for day, values in factors.items()
        },
    }
    temporary = target.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    temporary.replace(target)


def collect_factor_rows_by_day(
    companies: pd.DataFrame,
    history_outcomes: Mapping[str, HistoryOutcome],
    signal_dates: Sequence[date],
    *,
    cache_key: str,
    cache_hours: float,
) -> tuple[dict[date, list[dict[str, object]]], dict[date, dict[str, int]], pd.DataFrame]:
    """按交易日构造精算因子行，并在第二次运行时复用因子缓存。"""

    factors_by_day: dict[date, list[dict[str, object]]] = {day: [] for day in signal_dates}
    day_stats: dict[date, dict[str, int]] = {
        day: {
            "当日有行情股票数": 0,
            "长历史不可用股票数": 0,
            "历史预热不足股票数": 0,
            "基础预筛选通过数": 0,
            "策略预热不足候选数": 0,
            "精算因子行数": 0,
            "因子缓存命中数": 0,
            "因子计算失败数": 0,
        }
        for day in signal_dates
    }
    error_rows: list[dict[str, object]] = []

    for record in companies.to_dict("records"):
        code = str(record["股票代码"])
        outcome = history_outcomes.get(code)
        if outcome is None or outcome.history is None:
            for signal_day in signal_dates:
                day_stats[signal_day]["长历史不可用股票数"] += 1
            continue
        history = outcome.history
        if len(history) < strategy_app.MIN_REQUIRED_BARS:
            for signal_day in signal_dates:
                day_stats[signal_day]["历史预热不足股票数"] += 1
            continue
        available_dates = {pd.Timestamp(value).date() for value in history["date"]}
        for signal_day in signal_dates:
            if signal_day in available_dates:
                day_stats[signal_day]["当日有行情股票数"] += 1

        prefiltered_dates = _basic_prefilter_dates(history, signal_dates)
        for signal_day in prefiltered_dates:
            day_stats[signal_day]["基础预筛选通过数"] += 1
        if not prefiltered_dates:
            continue

        cached_factors = _read_factor_cache(
            code,
            cache_key=cache_key,
            history_token=outcome.cache_token,
            cache_hours=cache_hours,
        )
        cache_changed = False
        for signal_day in sorted(prefiltered_dates):
            date_key = signal_day.isoformat()
            history_bar_count = int(
                history["date"].searchsorted(pd.Timestamp(signal_day), side="right")
            )
            if history_bar_count < strategy_app.MIN_REQUIRED_BARS:
                day_stats[signal_day]["策略预热不足候选数"] += 1
                error_rows.append(
                    {
                        "序号": record["序号"],
                        "股票代码": code,
                        "股票名称": record["股票名称"],
                        "选股日期": signal_day,
                        "问题类型": "策略预热不足",
                        "失败原因": (
                            f"截至选股日仅有 {history_bar_count} 根有效日线，"
                            f"少于 KDJ 所需的 {strategy_app.MIN_REQUIRED_BARS} 根。"
                        ),
                    }
                )
                continue
            factor_values = cached_factors.get(date_key)
            if factor_values is None:
                try:
                    factor_values = strategy_app._calculate_factors_from_normalized_history(
                        history,
                        as_of_date=signal_day,
                    )
                    # 选股日停牌时原应用会退回前一根 K 线，历史回测不能把它当作当日信号。
                    if factor_values.get("数据日期") != date_key:
                        raise BacktestDataError("该股票在选股日没有可交易日线。")
                    cached_factors[date_key] = dict(factor_values)
                    cache_changed = True
                except (BacktestDataError, OSError, TypeError, ValueError) as exc:
                    day_stats[signal_day]["因子计算失败数"] += 1
                    error_rows.append(
                        {
                            "序号": record["序号"],
                            "股票代码": code,
                            "股票名称": record["股票名称"],
                            "选股日期": signal_day,
                            "问题类型": "精算因子失败",
                            "失败原因": str(exc),
                        }
                    )
                    continue
            else:
                day_stats[signal_day]["因子缓存命中数"] += 1

            if factor_values.get("数据日期") != date_key:
                day_stats[signal_day]["因子计算失败数"] += 1
                error_rows.append(
                    {
                        "序号": record["序号"],
                        "股票代码": code,
                        "股票名称": record["股票名称"],
                        "选股日期": signal_day,
                        "问题类型": "精算因子日期不匹配",
                        "失败原因": "缓存因子并非选股日数据，已忽略。",
                    }
                )
                continue
            factors_by_day[signal_day].append(
                {
                    "序号": record["序号"],
                    "股票代码": code,
                    "股票名称": record["股票名称"],
                    "所属行业": record.get("所属行业"),
                    "数据来源": outcome.source or "未知来源",
                    "缓存命中": outcome.from_cache,
                    **factor_values,
                }
            )
            day_stats[signal_day]["精算因子行数"] += 1

        if cache_changed and cache_hours > 0:
            _write_factor_cache(
                code,
                cached_factors,
                cache_key=cache_key,
                history_token=outcome.cache_token,
            )

    errors = pd.DataFrame(error_rows, columns=HISTORY_ERROR_COLUMNS)
    if not errors.empty:
        errors = errors.sort_values(["序号", "问题类型"], kind="stable").reset_index(drop=True)
    return factors_by_day, day_stats, errors


def _precompute_factor_bars(history: pd.DataFrame) -> pd.DataFrame:
    """一次性计算所有日期共享的技术指标时间序列。

    根应用的单日因子函数会按截至日重新计算整段滚动指标。历史回测需要大量日期时，
    这些指标都是只依赖过去数据的序列，可先按一只股票完整计算一次，再安全地按日取值。
    KDJ 健康金叉和近三日 K 线形态仍在具体日期上截断判断，以保持策略口径一致。
    """

    bars = history.copy()
    bars["ma5"] = bars["close"].rolling(5, min_periods=5).mean()
    bars["ma20"] = bars["close"].rolling(20, min_periods=20).mean()
    bars["volume_ma5"] = bars["volume"].shift(1).rolling(5, min_periods=5).mean()

    delta = bars["close"].diff()
    gains = delta.clip(lower=0)
    losses = -delta.clip(upper=0)
    average_gain = gains.ewm(alpha=1 / 14, adjust=False, min_periods=14).mean()
    average_loss = losses.ewm(alpha=1 / 14, adjust=False, min_periods=14).mean()
    relative_strength = average_gain / average_loss
    bars["rsi14"] = 100.0 - 100.0 / (1.0 + relative_strength)
    bars.loc[(average_loss == 0) & (average_gain > 0), "rsi14"] = 100.0
    bars.loc[(average_gain == 0) & (average_loss > 0), "rsi14"] = 0.0

    ema12 = bars["close"].ewm(span=12, adjust=False, min_periods=12).mean()
    ema26 = bars["close"].ewm(span=26, adjust=False, min_periods=26).mean()
    bars["macd_dif"] = ema12 - ema26
    bars["macd_dea"] = bars["macd_dif"].ewm(span=9, adjust=False, min_periods=9).mean()
    bars["macd_histogram"] = 2.0 * (bars["macd_dif"] - bars["macd_dea"])
    bars["macd_golden_cross"] = (
        bars["macd_dif"].gt(bars["macd_dea"])
        & bars["macd_dif"].shift(1).le(bars["macd_dea"].shift(1))
    ).fillna(False)
    bars["macd_dead_cross"] = (
        bars["macd_dif"].lt(bars["macd_dea"])
        & bars["macd_dif"].shift(1).ge(bars["macd_dea"].shift(1))
    ).fillna(False)

    bars["kdj_rsv"], bars["kdj_k"], bars["kdj_d"], bars["kdj_j"] = (
        strategy_app._calculate_kdj_series(bars)
    )
    bars["kdj_golden_cross"] = (
        bars["kdj_k"].gt(bars["kdj_d"])
        & bars["kdj_k"].shift(1).le(bars["kdj_d"].shift(1))
    ).fillna(False)
    bars["kdj_dead_cross"] = (
        bars["kdj_k"].lt(bars["kdj_d"])
        & bars["kdj_k"].shift(1).ge(bars["kdj_d"].shift(1))
    ).fillna(False)
    bars["prior_platform_high"] = (
        bars["high"]
        .shift(1)
        .rolling(
            strategy_app.PLATFORM_BREAKOUT_LOOKBACK_BARS,
            min_periods=strategy_app.PLATFORM_BREAKOUT_LOOKBACK_BARS,
        )
        .max()
    )
    bars["rolling_high_60"] = bars["high"].shift(1).rolling(60, min_periods=60).max()

    body = (bars["close"] - bars["open"]).abs()
    intraday_range = bars["high"] - bars["low"]
    body_top = pd.concat([bars["open"], bars["close"]], axis=1).max(axis=1)
    body_bottom = pd.concat([bars["open"], bars["close"]], axis=1).min(axis=1)
    upper_shadow = bars["high"] - body_top
    lower_shadow = body_bottom - bars["low"]
    valid_candle = (
        intraday_range.gt(0)
        & bars["high"].ge(body_top)
        & bars["low"].le(body_bottom)
        & upper_shadow.ge(0)
        & lower_shadow.ge(0)
    ).fillna(False)
    valid_range = intraday_range.where(valid_candle)
    body_ratio = body.div(valid_range)
    upper_shadow_ratio = upper_shadow.div(valid_range)
    lower_shadow_ratio = lower_shadow.div(valid_range)
    doji = valid_candle & body_ratio.le(
        strategy_app.CANDLESTICK_DOJI_MAX_BODY_RATIO
        + strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
    ).fillna(False)
    inverted_t_doji = (
        doji
        & lower_shadow_ratio.le(
            strategy_app.CANDLESTICK_INVERTED_T_MAX_LOWER_SHADOW_RATIO
            + strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        ).fillna(False)
        & upper_shadow_ratio.ge(
            strategy_app.CANDLESTICK_INVERTED_T_MIN_UPPER_SHADOW_RATIO
            - strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        ).fillna(False)
    )
    prior_high = bars["high"].shift(1).rolling(
        strategy_app.CANDLESTICK_HANGING_MAN_LOOKBACK_BARS,
        min_periods=strategy_app.CANDLESTICK_HANGING_MAN_LOOKBACK_BARS,
    ).max()
    hanging_man = (
        valid_candle
        & prior_high.notna()
        & bars["high"].ge(
            prior_high * strategy_app.CANDLESTICK_HANGING_MAN_HIGH_ZONE_RATIO
            - strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        )
        & body_ratio.le(
            strategy_app.CANDLESTICK_HANGING_MAN_MAX_BODY_RATIO
            + strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        ).fillna(False)
        & lower_shadow_ratio.ge(
            strategy_app.CANDLESTICK_HANGING_MAN_MIN_LOWER_SHADOW_RATIO
            - strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        ).fillna(False)
        & upper_shadow_ratio.le(
            strategy_app.CANDLESTICK_HANGING_MAN_MAX_UPPER_SHADOW_RATIO
            + strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        ).fillna(False)
    )
    long_upper_shadow_bullish = (
        valid_candle
        & bars["close"].gt(bars["open"])
        & upper_shadow.gt(body + strategy_app.CANDLESTICK_COMPARISON_TOLERANCE)
    ).fillna(False)
    daily_close_change = bars["close"].pct_change(fill_method=None).mul(100.0)
    extreme_bullish = (
        valid_candle
        & bars["close"].gt(bars["open"])
        & daily_close_change.gt(
            strategy_app.CANDLESTICK_EXTREME_BULLISH_MIN_PCT_CHANGE
            + strategy_app.CANDLESTICK_COMPARISON_TOLERANCE
        ).fillna(False)
    )
    bars["risk_doji_3d"] = doji
    bars["risk_inverted_t_doji_3d"] = inverted_t_doji
    bars["risk_hanging_man_3d"] = hanging_man
    bars["risk_long_upper_shadow_bullish_3d"] = long_upper_shadow_bullish
    bars["risk_extreme_bullish_3d"] = extreme_bullish
    return bars


def _factor_values_from_precomputed_bars(
    bars: pd.DataFrame,
    position: int,
) -> dict[str, object]:
    """从预计算序列构造一个交易日的因子，字段和值与根应用保持一致。"""

    latest = bars.iloc[position]
    previous = bars.iloc[position - 1]
    if pd.isna(latest["ma20"]) or pd.isna(latest["rsi14"]) or pd.isna(latest["macd_dea"]):
        raise ValueError("技术指标预热不足。")

    prior_platform_high = latest["prior_platform_high"]
    platform_breakout = bool(
        pd.notna(prior_platform_high)
        and pd.notna(latest["close"])
        and latest["close"] > prior_platform_high
    )
    platform_breakout_pct = (
        (latest["close"] - prior_platform_high) / prior_platform_high * 100.0
        if pd.notna(prior_platform_high) and prior_platform_high > 0
        else None
    )
    ma5_rising = bool(latest["ma5"] > previous["ma5"])
    volume_ratio = latest["volume"] / latest["volume_ma5"] if latest["volume_ma5"] > 0 else pd.NA
    macd_bullish = bool(latest["macd_dif"] > latest["macd_dea"])
    macd_bearish = bool(latest["macd_dif"] < latest["macd_dea"])
    macd_golden_cross = bool(latest["macd_golden_cross"])
    macd_dead_cross = bool(latest["macd_dead_cross"])

    macd_golden_cross_date: str | None = None
    macd_golden_cross_age: int | None = None
    completed_macd_position = position - strategy_app.MACD_GOLDEN_CROSS_OFFSET_BARS
    minimum_macd_position = max(
        1,
        completed_macd_position
        - int(strategy_app.MAX_MACD_GOLDEN_CROSS_LOOKBACK_DAYS)
        + 1,
    )
    has_macd_dead_cross_after = False
    for cross_position in range(
        completed_macd_position,
        minimum_macd_position - 1,
        -1,
    ):
        cross = bars.iloc[cross_position]
        if bool(cross["macd_dead_cross"]):
            has_macd_dead_cross_after = True
            continue
        if has_macd_dead_cross_after or not bool(cross["macd_golden_cross"]):
            continue
        macd_golden_cross_date = pd.Timestamp(cross["date"]).date().isoformat()
        macd_golden_cross_age = completed_macd_position - cross_position + 1
        break

    # 考察截至当前信号日的交易日。向前扫描时，一旦遇到死叉，更早的
    # 金叉均已失效；因此第一个仍有效的健康金叉就是最近的有效信号。
    kdj_healthy_golden_cross_date: str | None = None
    kdj_healthy_golden_cross_age: int | None = None
    completed_position = position - strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_OFFSET_BARS
    minimum_position = max(
        1,
        completed_position
        - int(strategy_app.MAX_KDJ_HEALTHY_GOLDEN_CROSS_AGE),
    )
    has_kdj_dead_cross_after = False
    for cross_position in range(completed_position, minimum_position - 1, -1):
        cross = bars.iloc[cross_position]
        if bool(cross["kdj_dead_cross"]):
            has_kdj_dead_cross_after = True
            continue
        if has_kdj_dead_cross_after or not bool(cross["kdj_golden_cross"]):
            continue
        prior_cross = bars.iloc[cross_position - 1]
        kdj_golden_cross_in_oversold = bool(
            cross["kdj_k"] < strategy_app.KDJ_OVERSOLD_THRESHOLD
            and cross["kdj_d"] < strategy_app.KDJ_OVERSOLD_THRESHOLD
            and cross["kdj_j"] < strategy_app.KDJ_OVERSOLD_THRESHOLD
        )
        kdj_j_rising = bool(
            pd.notna(cross["kdj_j"])
            and pd.notna(prior_cross["kdj_j"])
            and cross["kdj_j"] > prior_cross["kdj_j"]
        )
        kdj_top_divergence = strategy_app._has_kdj_top_divergence(
            bars.iloc[: cross_position + 1]
        )
        if (
            kdj_golden_cross_in_oversold
            and kdj_j_rising
            and not kdj_top_divergence
        ):
            kdj_healthy_golden_cross_date = pd.Timestamp(cross["date"]).date().isoformat()
            kdj_healthy_golden_cross_age = completed_position - cross_position
            break

    def numeric_or_none(value: object) -> float | None:
        return None if pd.isna(value) else float(value)

    float_market_cap_yi = strategy_app.estimated_float_market_cap_yi(
        latest["amount"], latest["turnover"]
    )
    bias20 = (latest["close"] - latest["ma20"]) / latest["ma20"] * 100.0
    intraday_range = latest["high"] - latest["low"]
    body_top = max(latest["open"], latest["close"])
    upper_shadow = max(0.0, latest["high"] - body_top)
    upper_shadow_ratio = (
        0.0
        if intraday_range <= 0
        else upper_shadow / intraday_range * 100.0
    )
    if pd.isna(intraday_range) or pd.isna(latest["close"]) or pd.isna(latest["low"]):
        close_position = None
    elif intraday_range > 0:
        close_position = (latest["close"] - latest["low"]) / intraday_range * 100.0
    elif latest["close"] == latest["high"]:
        close_position = 100.0
    else:
        close_position = None
    close_near_daily_high = bool(
        close_position is not None
        and close_position >= strategy_app.CLOSE_NEAR_DAILY_HIGH_THRESHOLD
    )
    rolling_high_60 = latest["rolling_high_60"]
    touches_60_day_resistance = bool(latest["close"] >= rolling_high_60 * 0.98)
    recent_risk_bars = bars.iloc[
        max(0, position - strategy_app.CANDLESTICK_RISK_LOOKBACK_BARS + 1) : position + 1
    ]
    candlestick_risk_flags = {
        "doji_3d": bool(recent_risk_bars["risk_doji_3d"].any()),
        "inverted_t_doji_3d": bool(recent_risk_bars["risk_inverted_t_doji_3d"].any()),
        "hanging_man_3d": bool(recent_risk_bars["risk_hanging_man_3d"].any()),
        "long_upper_shadow_bullish_3d": bool(
            recent_risk_bars["risk_long_upper_shadow_bullish_3d"].any()
        ),
        "extreme_bullish_3d": bool(recent_risk_bars["risk_extreme_bullish_3d"].any()),
    }

    return {
        "数据日期": pd.Timestamp(latest["date"]).date().isoformat(),
        "收盘价": numeric_or_none(latest["close"]),
        "MA5": numeric_or_none(latest["ma5"]),
        "MA20": numeric_or_none(latest["ma20"]),
        "站上MA5": bool(latest["close"] > latest["ma5"]),
        "站上MA20": bool(latest["close"] > latest["ma20"]),
        "MA5高于MA20": bool(latest["ma5"] > latest["ma20"]),
        "MA5上行": ma5_rising,
        "前20日平台最高价": numeric_or_none(prior_platform_high),
        "平台突破幅度（%）": numeric_or_none(platform_breakout_pct),
        "收盘突破前20日平台": platform_breakout,
        "RSI14": numeric_or_none(latest["rsi14"]),
        "MACD_DIF": numeric_or_none(latest["macd_dif"]),
        "MACD_DEA": numeric_or_none(latest["macd_dea"]),
        "MACD柱": numeric_or_none(latest["macd_histogram"]),
        "MACD多头": macd_bullish,
        "MACD空头": macd_bearish,
        "MACD金叉": macd_golden_cross,
        "MACD死叉": macd_dead_cross,
        strategy_app.MACD_GOLDEN_CROSS_DATE_COLUMN: macd_golden_cross_date,
        strategy_app.MACD_GOLDEN_CROSS_AGE_COLUMN: macd_golden_cross_age,
        "KDJ_K(89,3,3)": numeric_or_none(latest["kdj_k"]),
        "KDJ_D(89,3,3)": numeric_or_none(latest["kdj_d"]),
        "KDJ_J(89,3,3)": numeric_or_none(latest["kdj_j"]),
        "KDJ金叉": bool(latest["kdj_golden_cross"]),
        "KDJ死叉": bool(latest["kdj_dead_cross"]),
        strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_DATE_COLUMN: kdj_healthy_golden_cross_date,
        strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_AGE_COLUMN: kdj_healthy_golden_cross_age,
        strategy_app.KDJ_HEALTHY_GOLDEN_CROSS_NAME: (
            kdj_healthy_golden_cross_age is not None
        ),
        "当日成交量": numeric_or_none(latest["volume"]),
        "5日均量": numeric_or_none(latest["volume_ma5"]),
        "量比": numeric_or_none(volume_ratio),
        "放量": bool(
            not pd.isna(volume_ratio)
            and strategy_app.DEFAULT_VOLUME_RATIO_RANGE[0]
            <= volume_ratio
            <= strategy_app.DEFAULT_VOLUME_RATIO_RANGE[1]
        ),
        "当日成交额": numeric_or_none(latest["amount"]),
        "换手率": numeric_or_none(latest["turnover"]),
        "估算流通市值（亿元）": float_market_cap_yi,
        "当日涨跌幅": numeric_or_none(latest["pct_change"]),
        "振幅": numeric_or_none(latest["amplitude"]),
        "BIAS20": numeric_or_none(bias20),
        "上影线比例": numeric_or_none(upper_shadow_ratio),
        "收盘日内位置（%）": numeric_or_none(close_position),
        "收盘位于日内高位": close_near_daily_high,
        "60日最高价": numeric_or_none(rolling_high_60),
        "触及60日高点压力": touches_60_day_resistance,
        **{
            strategy_app.CANDLESTICK_RISK_FACTOR_COLUMNS[key]: hit
            for key, hit in candlestick_risk_flags.items()
        },
    }


def _calculate_factor_values_for_dates(
    history: pd.DataFrame,
    signal_dates: Sequence[date],
) -> dict[str, dict[str, object]]:
    """为一只股票的多个实际交易日生成与页面完全一致的完整因子。

    RSI、MACD 和 KDJ 都有递推初值。原页面的语义是每个截至日仅保留最近
    ``INDICATOR_WARMUP_BARS`` 根日线，再从该窗口开头重新计算。因此不能把
    640 根长历史预计算成一条连续指标序列后直接取最后一行；那会把窗口外的
    价格带入递推状态。这里显式构造同一窗口并委托页面的唯一因子实现，确保
    回测和浏览器筛选逐字段一致。
    """

    return factor_batch.calculate_factor_values_for_dates(
        history,
        signal_dates,
        strategy_app=strategy_app,
    )


def _safe_calculate_factor_values_for_dates(
    history: pd.DataFrame,
    signal_dates: Sequence[date],
) -> tuple[dict[str, dict[str, object]], str | None]:
    """供并发任务调用的因子计算包装，避免单只异常中止整批回测。"""

    try:
        return _calculate_factor_values_for_dates(history, signal_dates), None
    except (OSError, TypeError, ValueError) as exc:
        return {}, str(exc)
    except Exception as exc:  # pragma: no cover - 并发任务异常的防御性兜底。
        return {}, f"因子任务异常：{exc}"


def _precalculate_missing_full_factor_values(
    records: Sequence[Mapping[str, object]],
    history_outcomes: Mapping[str, HistoryOutcome],
    signal_dates: Sequence[date],
    *,
    cache_key: str,
    cache_hours: float,
    factor_workers: int,
    progress_callback: HistoryProgressCallback | None,
) -> dict[str, tuple[dict[str, dict[str, object]], str | None]]:
    """按股票并行预计算未命中缓存的完整因子，磁盘写入仍由主进程完成。"""

    full_factor_cache_key = _full_factor_cache_key(cache_key)
    ordered_dates = tuple(sorted(set(signal_dates)))
    jobs: list[tuple[str, pd.DataFrame, tuple[date, ...]]] = []
    cached_factor_count = 0
    for record in records:
        code = str(record["股票代码"])
        outcome = history_outcomes.get(code)
        if outcome is None or outcome.history is None:
            continue
        history = outcome.history
        if len(history) < strategy_app.MIN_REQUIRED_BARS:
            continue
        positions = {
            pd.Timestamp(value).date(): position
            for position, value in enumerate(history["date"])
        }
        factor_dates = tuple(
            signal_day
            for signal_day in ordered_dates
            if (position := positions.get(signal_day)) is not None
            and position + 1 >= strategy_app.MIN_REQUIRED_BARS
        )
        if not factor_dates:
            continue
        cached_factors = _read_factor_cache(
            code,
            cache_key=full_factor_cache_key,
            history_token=outcome.cache_token,
            cache_hours=cache_hours,
            cache_version=FULL_FACTOR_CACHE_VERSION,
        )
        missing_dates = tuple(
            signal_day
            for signal_day in factor_dates
            if not (
                isinstance(cached_value := cached_factors.get(signal_day.isoformat()), Mapping)
                and cached_value.get("数据日期") == signal_day.isoformat()
            )
        )
        cached_factor_count += len(factor_dates) - len(missing_dates)
        if missing_dates:
            jobs.append((code, history, missing_dates))

    if not jobs or factor_workers <= 1:
        return {}

    values_by_code: dict[str, tuple[dict[str, dict[str, object]], str | None]] = {}
    completed = 0
    failed = 0
    base_completed = len(records) - len(jobs)
    try:
        # Streamlit 在 Windows 上运行时，进程池会按 spawn 方式重新执行
        # backtest_app.py 的顶层页面代码。因子计算主要是 pandas/numpy 操作，
        # 用线程池既可并发，又不会触发页面重入。
        with ThreadPoolExecutor(max_workers=min(factor_workers, len(jobs))) as executor:
            futures = {
                executor.submit(_safe_calculate_factor_values_for_dates, history, dates): code
                for code, history, dates in jobs
            }
            for future in as_completed(futures):
                code = futures[future]
                try:
                    values, error = future.result()
                except Exception as exc:  # pragma: no cover - 并发任务异常的防御性兜底。
                    values, error = {}, f"因子任务异常：{exc}"
                values_by_code[code] = (values, error)
                completed += 1
                if error is not None:
                    failed += 1
                if progress_callback is not None and (
                    completed == len(jobs) or completed % 10 == 0
                ):
                    progress_callback(
                        base_completed + completed,
                        len(records),
                        code,
                        cached_factor_count,
                        base_completed + completed - failed,
                        failed,
                    )
    except Exception:
        # 并发任务无法启动时，主循环会仅对尚未得到结果的股票走相同的
        # 逐窗口黄金路径，正确性优先于并发能力。
        return values_by_code
    return values_by_code


def collect_all_factor_rows_by_day(
    companies: pd.DataFrame,
    history_outcomes: Mapping[str, HistoryOutcome],
    signal_dates: Sequence[date],
    *,
    cache_key: str,
    cache_hours: float,
    progress_callback: HistoryProgressCallback | None = None,
    factor_workers: int | None = None,
) -> tuple[dict[date, list[dict[str, object]]], dict[date, dict[str, int]], pd.DataFrame]:
    """构造选定交易日的全量因子面板，不使用任何固定策略的预筛选。

    因子结果按股票和交易日写入现有磁盘缓存；随后更改勾选条件只需本地筛选，
    不会遗漏先前固定预筛没有覆盖的候选股票。
    """

    ordered_dates = tuple(sorted(set(signal_dates)))
    full_factor_cache_key = _full_factor_cache_key(cache_key)
    factors_by_day: dict[date, list[dict[str, object]]] = {day: [] for day in ordered_dates}
    day_stats: dict[date, dict[str, int]] = {
        day: {
            "当日有行情股票数": 0,
            "长历史不可用股票数": 0,
            "历史预热不足股票数": 0,
            "因子预热不足股票数": 0,
            "精算因子行数": 0,
            "因子缓存命中数": 0,
            "因子计算失败数": 0,
        }
        for day in ordered_dates
    }
    error_rows: list[dict[str, object]] = []
    records = companies.to_dict("records")
    requested_factor_workers = (
        DEFAULT_FACTOR_WORKERS if factor_workers is None else int(factor_workers)
    )
    requested_factor_workers = max(
        1,
        min(requested_factor_workers, MAX_FACTOR_WORKERS),
    )
    precalculated_by_code = _precalculate_missing_full_factor_values(
        records,
        history_outcomes,
        ordered_dates,
        cache_key=cache_key,
        cache_hours=cache_hours,
        factor_workers=requested_factor_workers,
        progress_callback=progress_callback,
    )
    completed = cache_hits = succeeded = failed = 0

    for record in records:
        code = str(record["股票代码"])
        outcome = history_outcomes.get(code)
        completed += 1
        if outcome is None or outcome.history is None:
            for signal_day in ordered_dates:
                day_stats[signal_day]["长历史不可用股票数"] += 1
            if progress_callback is not None and (
                not precalculated_by_code or completed == len(records)
            ):
                progress_callback(completed, len(records), code, cache_hits, succeeded, failed)
            continue

        history = outcome.history
        if len(history) < strategy_app.MIN_REQUIRED_BARS:
            for signal_day in ordered_dates:
                day_stats[signal_day]["历史预热不足股票数"] += 1
            if progress_callback is not None and (
                not precalculated_by_code or completed == len(records)
            ):
                progress_callback(completed, len(records), code, cache_hits, succeeded, failed)
            continue
        date_positions = {
            pd.Timestamp(value).date(): position
            for position, value in enumerate(history["date"])
        }
        factor_dates: list[date] = []
        for signal_day in ordered_dates:
            position = date_positions.get(signal_day)
            if position is None:
                continue
            day_stats[signal_day]["当日有行情股票数"] += 1
            if position + 1 < strategy_app.MIN_REQUIRED_BARS:
                day_stats[signal_day]["因子预热不足股票数"] += 1
                error_rows.append(
                    {
                        "序号": record["序号"],
                        "股票代码": code,
                        "股票名称": record["股票名称"],
                        "选股日期": signal_day,
                        "问题类型": "因子预热不足",
                        "失败原因": (
                            f"截至选股日仅有 {position + 1} 根有效日线，"
                            f"少于 KDJ 所需的 {strategy_app.MIN_REQUIRED_BARS} 根。"
                        ),
                    }
                )
                continue
            factor_dates.append(signal_day)

        cached_factors = _read_factor_cache(
            code,
            cache_key=full_factor_cache_key,
            history_token=outcome.cache_token,
            cache_hours=cache_hours,
            cache_version=FULL_FACTOR_CACHE_VERSION,
        )
        valid_values: dict[str, dict[str, object]] = {}
        missing_dates: list[date] = []
        for signal_day in factor_dates:
            date_key = signal_day.isoformat()
            cached_value = cached_factors.get(date_key)
            if isinstance(cached_value, Mapping) and cached_value.get("数据日期") == date_key:
                valid_values[date_key] = dict(cached_value)
                day_stats[signal_day]["因子缓存命中数"] += 1
                cache_hits += 1
            else:
                missing_dates.append(signal_day)

        cache_changed = False
        if missing_dates:
            precalculated = precalculated_by_code.get(code)
            if precalculated is not None:
                calculated_values, calculation_error = precalculated
            else:
                calculated_values, calculation_error = _safe_calculate_factor_values_for_dates(
                    history,
                    missing_dates,
                )
            for signal_day in missing_dates:
                date_key = signal_day.isoformat()
                factor_values = calculated_values.get(date_key)
                if factor_values is None:
                    try:
                        # 防御性回退仍显式截取同一预热窗口，不能让窗口外的
                        # 递推状态进入 RSI、MACD 或 KDJ。
                        position = date_positions[signal_day]
                        window_start = max(
                            0,
                            position - strategy_app.INDICATOR_WARMUP_BARS + 1,
                        )
                        window = history.iloc[window_start : position + 1].reset_index(
                            drop=True
                        )
                        factor_values = strategy_app._calculate_factors_from_normalized_history(
                            window,
                            as_of_date=signal_day,
                        )
                    except (OSError, TypeError, ValueError) as exc:
                        day_stats[signal_day]["因子计算失败数"] += 1
                        failed += 1
                        error_rows.append(
                            {
                                "序号": record["序号"],
                                "股票代码": code,
                                "股票名称": record["股票名称"],
                                "选股日期": signal_day,
                                "问题类型": "精算因子失败",
                                "失败原因": calculation_error or str(exc),
                            }
                        )
                        continue
                if factor_values.get("数据日期") != date_key:
                    day_stats[signal_day]["因子计算失败数"] += 1
                    failed += 1
                    error_rows.append(
                        {
                            "序号": record["序号"],
                            "股票代码": code,
                            "股票名称": record["股票名称"],
                            "选股日期": signal_day,
                            "问题类型": "精算因子日期不匹配",
                            "失败原因": "计算结果并非选股日数据，已忽略。",
                        }
                    )
                    continue
                factor_copy = dict(factor_values)
                valid_values[date_key] = factor_copy
                cached_factors[date_key] = factor_copy
                cache_changed = True

        for signal_day in factor_dates:
            factor_values = valid_values.get(signal_day.isoformat())
            if factor_values is None:
                continue
            factors_by_day[signal_day].append(
                {
                    "序号": record["序号"],
                    "股票代码": code,
                    "股票名称": record["股票名称"],
                    "所属行业": record.get("所属行业"),
                    "数据来源": outcome.source or "未知来源",
                    "缓存命中": outcome.from_cache,
                    **factor_values,
                }
            )
            day_stats[signal_day]["精算因子行数"] += 1

        if cache_changed and cache_hours > 0:
            _write_factor_cache(
                code,
                cached_factors,
                cache_key=full_factor_cache_key,
                history_token=outcome.cache_token,
                cache_version=FULL_FACTOR_CACHE_VERSION,
            )
        succeeded += 1
        if progress_callback is not None and (
            completed == len(records)
            or (not precalculated_by_code and completed % 10 == 0)
        ):
            progress_callback(completed, len(records), code, cache_hits, succeeded, failed)

    errors = pd.DataFrame(error_rows, columns=HISTORY_ERROR_COLUMNS)
    if not errors.empty:
        errors = errors.sort_values(
            ["选股日期", "序号", "问题类型"],
            kind="stable",
            na_position="last",
        ).reset_index(drop=True)
    return factors_by_day, day_stats, errors


def _selection_reason(
    *,
    exact_factor_count: int,
    qualifying_count: int,
    risk_excluded_count: int,
    require_all: bool,
) -> str:
    if exact_factor_count == 0:
        return "无股票通过基础预筛选"
    if qualifying_count == 0:
        return "无股票满足全部指标条件" if require_all else "无股票满足任一指标条件"
    if risk_excluded_count >= qualifying_count:
        return "满足指标的股票均被风险过滤剔除"
    return "无可用候选"


def _attach_industries_to_ranked_candidates(
    ranked_candidates: pd.DataFrame,
    factors: pd.DataFrame,
) -> pd.DataFrame:
    """把股票池行业附回评分函数返回的有序候选，且不改变既有排序。"""

    if (
        ranked_candidates.empty
        or "股票代码" not in ranked_candidates.columns
        or "股票代码" not in factors.columns
        or "所属行业" not in factors.columns
    ):
        return ranked_candidates
    industries = factors.loc[:, ["股票代码", "所属行业"]].copy()
    industries["股票代码"] = industries["股票代码"].astype(str)
    industry_by_code = (
        industries.drop_duplicates(subset=["股票代码"], keep="first")
        .set_index("股票代码")["所属行业"]
    )
    result = ranked_candidates.copy()
    result["所属行业"] = result["股票代码"].astype(str).map(industry_by_code)
    return result


def evaluate_strategy(
    return_data: ReturnData,
    factors_by_day: Mapping[date, Sequence[Mapping[str, object]]],
    day_stats: Mapping[date, Mapping[str, int]],
    *,
    selected: Mapping[str, bool],
    selected_risks: Mapping[str, bool],
    turnover_range: tuple[float, float],
    float_market_cap_range_yi: tuple[float, float],
    pct_change_range: tuple[float, float],
    amplitude_threshold: float,
    rsi_range: tuple[float, float] = strategy_app.DEFAULT_RSI_RANGE,
    macd_golden_cross_lookback_days: int = (
        strategy_app.DEFAULT_MACD_GOLDEN_CROSS_LOOKBACK_DAYS
    ),
    require_all: bool,
    kdj_healthy_golden_cross_age_range: tuple[int, int] = (
        strategy_app.DEFAULT_KDJ_HEALTHY_GOLDEN_CROSS_AGE_RANGE
    ),
    macd_dea_minus_dif_range: tuple[float, float] = (
        strategy_app.DEFAULT_MACD_DEA_MINUS_DIF_RANGE
    ),
    volume_ratio_range: tuple[float, float] = strategy_app.DEFAULT_VOLUME_RATIO_RANGE,
    volume_ratio_threshold: float | None = None,
    top_n: int = 1,
    realized_returns: Mapping[tuple[date, str], float] | None = None,
) -> tuple[pd.DataFrame, dict[str, object]]:
    """按风险过滤后的评分第一名选出一股，并计算严格次日收益。"""

    strategy_app.validate_selected_conditions(
        selected,
        require_all=require_all,
        macd_dea_minus_dif_range=macd_dea_minus_dif_range,
    )
    macd_lookback_days = strategy_app._macd_golden_cross_lookback_days(
        macd_golden_cross_lookback_days
    )
    volume_range = strategy_app._volume_ratio_range_bounds(volume_ratio_range)
    volume_threshold = (
        strategy_app._volume_ratio_threshold(volume_ratio_threshold)
        if volume_ratio_threshold is not None
        else None
    )
    kdj_age_range = strategy_app._kdj_healthy_golden_cross_age_range(
        kdj_healthy_golden_cross_age_range
    )
    candidate_limit = max(1, int(top_n))
    max_score = float(strategy_app.maximum_score(selected))
    return_lookup = return_data.strict_returns if realized_returns is None else realized_returns
    cumulative_multiplier = 1.0
    daily_rows: list[dict[str, object]] = []
    correct_count = 0
    prediction_count = 0
    failed_prediction_count = 0
    signal_count = 0
    no_signal_count = 0
    missing_return_count = 0
    data_incomplete_count = 0

    for signal_day in return_data.signal_dates:
        next_day = return_data.next_trade_dates[signal_day]
        raw_rows = list(factors_by_day.get(signal_day, ()))
        stats = dict(day_stats.get(signal_day, {}))
        all_conditions_count = 0
        qualifying_count = 0
        risk_excluded_count = 0
        eligible_count = 0
        selection = pd.DataFrame()
        # 长历史、预热不足和单股精算失败都会在因子面板前按股票剔除。只有
        # 当天没有任何可用因子行时，才不能形成可评估预测。
        data_incomplete = (
            int(stats.get("因子计算失败数", 0)) > 0 and not raw_rows
        )
        if raw_rows:
            factors = pd.DataFrame(raw_rows)
            matrix = strategy_app.condition_matrix(
                factors,
                selected,
                turnover_range=turnover_range,
                float_market_cap_range_yi=float_market_cap_range_yi,
                pct_change_range=pct_change_range,
                amplitude_threshold=amplitude_threshold,
                rsi_range=rsi_range,
                macd_golden_cross_lookback_days=macd_lookback_days,
                kdj_healthy_golden_cross_age_range=kdj_age_range,
                macd_dea_minus_dif_range=macd_dea_minus_dif_range,
                volume_ratio_range=volume_range,
                volume_ratio_threshold=volume_threshold,
            )
            passed_all = matrix.all(axis=1)
            all_conditions_count = int(passed_all.sum())
            qualifying = passed_all if require_all else matrix.any(axis=1)
            qualifying_count = int(qualifying.sum())
            selection, eligible_count, risk_excluded_count = strategy_app.score_and_select(
                factors,
                selected,
                selected_risks=selected_risks,
                turnover_range=turnover_range,
                float_market_cap_range_yi=float_market_cap_range_yi,
                pct_change_range=pct_change_range,
                amplitude_threshold=amplitude_threshold,
                rsi_range=rsi_range,
                macd_golden_cross_lookback_days=macd_lookback_days,
                kdj_healthy_golden_cross_age_range=kdj_age_range,
                macd_dea_minus_dif_range=macd_dea_minus_dif_range,
                volume_ratio_range=volume_range,
                volume_ratio_threshold=volume_threshold,
                require_all=require_all,
                top_n=candidate_limit,
            )
            selection = _attach_industries_to_ranked_candidates(selection, factors)

        row: dict[str, object] = {
            "选股日期": signal_day,
            "下一市场交易日": next_day,
            **stats,
            "完整满足指标数": all_conditions_count,
            "风险剔除数": risk_excluded_count,
            "最终候选数": eligible_count,
            "策略满分": max_score,
            "选中股票代码": None,
            "选中股票名称": None,
            "选中所属行业": None,
            "选中得分": None,
            "选中数据日期": None,
            "选中数据来源": None,
            "满足条件": None,
            "未满足条件（扣分项）": None,
            "次日真实涨跌幅（%）": None,
            "是否预测正确": "未预测",
            "当日组合收益率（%）": 0.0,
            "数据覆盖状态": "不完整" if data_incomplete else "完整",
            "状态": None,
            "说明": None,
        }
        if data_incomplete:
            data_incomplete_count += 1
            row["是否预测正确"] = "未预测"
            row["状态"] = "数据不完整"
            row["说明"] = (
                f"有 {int(stats.get('因子计算失败数', 0))} 条精算因子失败，"
                "且没有其余可用股票，本日未形成可评估预测，按 0% 收益计。"
            )
        elif selection.empty:
            no_signal_count += 1
            row["状态"] = "无信号"
            row["说明"] = _selection_reason(
                exact_factor_count=int(stats.get("精算因子行数", 0)),
                qualifying_count=qualifying_count,
                risk_excluded_count=risk_excluded_count,
                require_all=require_all,
            )
        else:
            signal_count += 1
            chosen = selection.iloc[0]
            code = str(chosen["股票代码"])
            row.update(
                {
                    "选中股票代码": code,
                    "选中股票名称": chosen.get("股票名称"),
                    "选中所属行业": chosen.get("所属行业"),
                    "选中得分": chosen.get("得分"),
                    "选中数据日期": chosen.get("数据日期"),
                    "选中数据来源": chosen.get("数据来源"),
                    "满足条件": chosen.get("满足条件"),
                    "未满足条件（扣分项）": chosen.get("未满足条件（扣分项）"),
                }
            )
            next_return = return_lookup.get((signal_day, code))
            if next_return is None:
                missing_return_count += 1
                row["状态"] = "次日收益缺失"
                row["说明"] = "该股票没有严格匹配的下一市场交易日收益，按 0% 计。"
            else:
                prediction_count += 1
                row["次日真实涨跌幅（%）"] = float(next_return)
                row["当日组合收益率（%）"] = float(next_return)
                if next_return > 0.0:
                    correct_count += 1
                    row["是否预测正确"] = "正确"
                    row["状态"] = "预测正确"
                else:
                    failed_prediction_count += 1
                    row["是否预测正确"] = "错误"
                    row["状态"] = "预测错误"
                row["说明"] = "使用严格下一市场交易日收益计算。"

        daily_return = float(row["当日组合收益率（%）"])
        cumulative_multiplier *= 1.0 + daily_return / 100.0
        row["累计收益率（%）"] = (cumulative_multiplier - 1.0) * 100.0
        daily_rows.append(row)

    daily_results = pd.DataFrame(daily_rows)
    total_days = len(return_data.signal_dates)
    unpredicted_count = total_days - prediction_count
    total_return = (cumulative_multiplier - 1.0) * 100.0
    summary: dict[str, object] = {
        "最终统计": f"{correct_count}/{prediction_count}，{total_return:.2f}%",
        "预测正确天数": correct_count,
        "预测天数": prediction_count,
        "失败预测天数": failed_prediction_count,
        "预测错误天数": failed_prediction_count,
        "未预测天数": unpredicted_count,
        "总天数": total_days,
        "预测正确率（%）": (
            correct_count / prediction_count * 100.0 if prediction_count else 0.0
        ),
        "总收益率（%）": total_return,
        "日历选股日数": len(return_data.signal_dates),
        "数据不完整天数": data_incomplete_count,
        "有信号天数": signal_count,
        "无信号天数": no_signal_count,
        "次日收益缺失天数": missing_return_count,
        "策略满分": max_score,
        "持仓规则": (
            "每个选股日仅买入一只股票：取风险过滤后的评分排名第一名；"
            "得分相同依次按量比、收盘日内位置、成交额降序，再按股票代码升序。"
        ),
        "收益规则": "严格下一市场交易日收盘相对选股日收盘的前复权涨跌幅；未预测日按 0% 计入复利。",
        "预测正确规则": "仅选中股票且存在严格下一交易日收益时形成预测；涨幅大于 0% 记为正确，其余预测记为失败。",
    }
    return daily_results, summary


def evaluate_fixed_strategy(
    return_data: ReturnData,
    factors_by_day: Mapping[date, Sequence[Mapping[str, object]]],
    day_stats: Mapping[date, Mapping[str, int]],
    *,
    realized_returns: Mapping[tuple[date, str], float] | None = None,
) -> tuple[pd.DataFrame, dict[str, object]]:
    """以原有固定参数调用通用回测评估函数。"""

    selected, selected_risks = fixed_strategy_settings()
    return evaluate_strategy(
        return_data,
        factors_by_day,
        day_stats,
        selected=selected,
        selected_risks=selected_risks,
        turnover_range=FIXED_TURNOVER_RANGE,
        float_market_cap_range_yi=FIXED_FLOAT_MARKET_CAP_RANGE_YI,
        pct_change_range=FIXED_PCT_CHANGE_RANGE,
        amplitude_threshold=FIXED_AMPLITUDE_THRESHOLD,
        require_all=True,
        top_n=1,
        realized_returns=realized_returns,
    )


def default_output_path(first_signal_date: date, last_signal_date: date) -> Path:
    """生成独立回测报表默认路径。"""

    return MODULE_DIR / "outputs" / (
        f"深市主板固定策略回测_{first_signal_date.isoformat()}_{last_signal_date.isoformat()}.xlsx"
    )


def _format_report_sheet(worksheet: Any, *, freeze_panes: str = "A2") -> None:
    """统一设置回测报表的基础可读性。"""

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[column_letter].width = min(max(width, 12), 42)
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for header in ("选股日期", "下一市场交易日"):
        if header in headers:
            for cell in worksheet[get_column_letter(headers[header])][1:]:
                cell.number_format = "yyyy-mm-dd"
    percent_headers = {
        "预测正确率（%）",
        "总收益率（%）",
        "次日真实涨跌幅（%）",
        "当日组合收益率（%）",
        "累计收益率（%）",
    }
    for header in percent_headers:
        if header in headers:
            for cell in worksheet[get_column_letter(headers[header])][1:]:
                cell.number_format = "0.00"
    if "股票代码" in headers:
        for cell in worksheet[get_column_letter(headers["股票代码"])][1:]:
            cell.number_format = "@"
    if "选中股票代码" in headers:
        for cell in worksheet[get_column_letter(headers["选中股票代码"])][1:]:
            cell.number_format = "@"


def write_backtest_workbook(
    daily_results: pd.DataFrame,
    summary: Mapping[str, object],
    data_problems: pd.DataFrame,
    output_path: Path,
) -> None:
    """输出汇总、每日回测和数据问题三张工作表。"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary_frame = pd.DataFrame(
        [{"指标": key, "数值": value} for key, value in summary.items()]
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_frame.to_excel(writer, sheet_name="汇总", index=False)
        daily_results.to_excel(writer, sheet_name="每日回测", index=False)
        data_problems.reindex(columns=HISTORY_ERROR_COLUMNS).to_excel(
            writer,
            sheet_name="数据问题",
            index=False,
        )
        _format_report_sheet(writer.sheets["汇总"])
        _format_report_sheet(writer.sheets["每日回测"], freeze_panes="C2")
        _format_report_sheet(writer.sheets["数据问题"])
