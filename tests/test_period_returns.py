import json
import unittest
from datetime import date, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import fetch_period_returns as app


def close_history(rows: list[tuple[str, float]]) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "date": pd.to_datetime([row[0] for row in rows]),
            "close": [row[1] for row in rows],
        }
    )


def outcome_from_history(
    code: str,
    history: pd.DataFrame,
    *,
    requested_start: date = date(2026, 7, 1),
    requested_end: date = date(2026, 7, 2),
    source: str = "test-source",
) -> app.PeriodReturnOutcome:
    return app._outcome_from_history(
        code,
        source,
        False,
        requested_start,
        requested_end,
        history,
    )


def row(columns: tuple[str, ...], values: list[object]) -> dict[str, object]:
    return dict(zip(columns, values, strict=True))


class PeriodReturnParsingTests(unittest.TestCase):
    def test_eastmoney_history_parses_adjusted_close_sorts_and_drops_invalid_rows(self) -> None:
        history = app.parse_eastmoney_history(
            {
                "data": {
                    "klines": [
                        "2026-07-03,10,11,12,9",
                        "invalid,10,12,12,9",
                        "2026-07-01,10,10,10,9",
                        "2026-07-03,10,12,12,9",
                    ]
                }
            }
        )

        self.assertEqual(history["date"].dt.date.tolist(), [date(2026, 7, 1), date(2026, 7, 3)])
        self.assertEqual(history["close"].tolist(), [10.0, 12.0])

    def test_tencent_rich_history_uses_qfq_close_not_unadjusted_day_close(self) -> None:
        payload = {
            "code": 0,
            "data": {
                "sz000001": {
                    "qfqday": [["2026-07-01", "9.8", "10.0", "10.1", "9.7", "1"]],
                    "day": [["2026-07-01", "19.8", "20.0", "20.1", "19.7", "1"]],
                }
            },
        }

        history = app.parse_tencent_rich_history(f"kline={json.dumps(payload)};", "sz000001")

        self.assertEqual(history.iloc[0]["close"], 10.0)

    def test_tencent_history_rejects_response_without_qfq_series(self) -> None:
        payload = {
            "code": 0,
            "data": {"sz000001": {"day": [["2026-07-01", "19.8", "20.0"]]}},
        }

        with self.assertRaises(app.PeriodReturnError):
            app.parse_tencent_rich_history(f"kline={json.dumps(payload)};", "sz000001")


class PeriodReturnCalculationTests(unittest.TestCase):
    def test_select_period_return_uses_only_trading_days_inside_requested_period(self) -> None:
        history = close_history(
            [
                ("2026-07-03", 10.0),
                ("2026-07-06", 11.0),
                ("2026-07-10", 12.5),
                ("2026-07-13", 13.0),
            ]
        )

        actual_start, start_close, actual_end, end_close, change_pct = app.select_period_return(
            history,
            date(2026, 7, 4),
            date(2026, 7, 11),
        )

        self.assertEqual(actual_start, date(2026, 7, 6))
        self.assertEqual(actual_end, date(2026, 7, 10))
        self.assertEqual(start_close, 11.0)
        self.assertEqual(end_close, 12.5)
        self.assertAlmostEqual(change_pct, (12.5 / 11.0 - 1.0) * 100.0)

    def test_select_period_return_rejects_data_outside_requested_period(self) -> None:
        history = close_history([("2026-07-03", 10.0), ("2026-07-13", 12.0)])

        with self.assertRaises(app.PeriodReturnError):
            app.select_period_return(history, date(2026, 7, 4), date(2026, 7, 11))

    def test_daily_returns_use_pre_period_close_for_first_in_period_trading_day(self) -> None:
        history = close_history(
            [
                ("2026-07-03", 10.0),
                ("2026-07-06", 11.0),
                ("2026-07-07", 10.45),
                ("2026-07-08", 11.495),
            ]
        )

        daily_returns = app.select_daily_returns(history, date(2026, 7, 4), date(2026, 7, 7))

        self.assertEqual([record.trade_date for record in daily_returns], [date(2026, 7, 6), date(2026, 7, 7)])
        self.assertEqual(daily_returns[0].previous_trade_date, date(2026, 7, 3))
        self.assertEqual(daily_returns[0].previous_close, 10.0)
        self.assertEqual(daily_returns[0].close, 11.0)
        self.assertAlmostEqual(daily_returns[0].change_pct or 0.0, 10.0)
        self.assertEqual(daily_returns[1].previous_trade_date, date(2026, 7, 6))
        self.assertEqual(daily_returns[1].previous_close, 11.0)
        self.assertAlmostEqual(daily_returns[1].change_pct or 0.0, -5.0)

    def test_daily_return_has_no_synthetic_change_without_previous_trading_day(self) -> None:
        history = close_history([("2026-07-06", 11.0)])

        daily_returns = app.select_daily_returns(history, date(2026, 7, 4), date(2026, 7, 7))

        self.assertEqual(len(daily_returns), 1)
        self.assertIsNone(daily_returns[0].previous_trade_date)
        self.assertIsNone(daily_returns[0].previous_close)
        self.assertIsNone(daily_returns[0].change_pct)

    def test_long_period_uses_contiguous_overlapping_windows(self) -> None:
        requested_start = date(2020, 1, 1)
        requested_end = date(2026, 7, 23)
        windows = app._query_windows(requested_start, requested_end)

        self.assertGreater(len(windows), 2)
        self.assertEqual(windows[0][0], requested_start)
        self.assertEqual(windows[-1][1], requested_end)
        self.assertTrue(
            all(
                (window_end - window_start).days + 1 <= app.MAX_SINGLE_QUERY_CALENDAR_DAYS
                for window_start, window_end in windows
            )
        )
        for (_, previous_end), (next_start, _) in zip(windows, windows[1:]):
            self.assertEqual(
                (previous_end - next_start).days + 1,
                app.QUERY_WINDOW_OVERLAP_DAYS,
            )

    def test_fetch_history_includes_lookback_needed_for_first_daily_return(self) -> None:
        requested_start = date(2026, 7, 1)
        requested_end = date(2026, 7, 2)
        history = close_history([("2026-06-30", 10.0), ("2026-07-01", 11.0)])
        windows = ((date(2026, 5, 31), requested_end),)

        with (
            patch.object(app, "_query_windows", return_value=windows) as query_windows,
            patch.object(app, "_fetch_windows", return_value=history) as fetch_windows,
        ):
            actual_history, source = app.fetch_period_history(
                "000001",
                requested_start,
                requested_end,
                app.RequestRateLimiter(0.0),
                app.SourceCircuitBreaker("eastmoney"),
                app.SourceCircuitBreaker("tencent"),
                1.0,
            )

        self.assertIs(actual_history, history)
        self.assertTrue(source)
        query_windows.assert_called_once_with(
            requested_start - timedelta(days=app.HISTORY_LOOKBACK_CALENDAR_DAYS),
            requested_end,
        )
        self.assertEqual(fetch_windows.call_args.args[2], windows)


class PeriodReturnCacheTests(unittest.TestCase):
    def test_cached_bars_rebuild_summary_and_each_daily_return(self) -> None:
        requested_start = date(2026, 7, 1)
        requested_end = date(2026, 7, 2)
        history = close_history(
            [
                ("2026-06-30", 10.0),
                ("2026-07-01", 11.0),
                ("2026-07-02", 10.45),
            ]
        )
        outcome = outcome_from_history(
            "000001",
            history,
            requested_start=requested_start,
            requested_end=requested_end,
        )

        with TemporaryDirectory() as temporary_directory:
            with patch.object(app, "CACHE_DIR", Path(temporary_directory)):
                app._write_cached_outcome(outcome)
                cache_path = app._cache_path("000001", requested_start, requested_end)
                payload = json.loads(cache_path.read_text(encoding="utf-8"))
                cached = app._read_cached_outcome(
                    "000001", requested_start, requested_end, cache_hours=1.0
                )

        self.assertEqual(payload["version"], app.CACHE_VERSION)
        self.assertEqual(payload["bars"], [
            {"date": "2026-06-30", "close": 10.0},
            {"date": "2026-07-01", "close": 11.0},
            {"date": "2026-07-02", "close": 10.45},
        ])
        self.assertIsNotNone(cached)
        assert cached is not None
        self.assertTrue(cached.from_cache)
        self.assertAlmostEqual(cached.change_pct, 10.45 / 11.0 * 100.0 - 100.0)
        self.assertEqual([record.trade_date for record in cached.daily_returns], [requested_start, requested_end])
        self.assertEqual(cached.daily_returns[0].previous_trade_date, date(2026, 6, 30))
        self.assertAlmostEqual(cached.daily_returns[0].change_pct or 0.0, 10.0)

    def test_legacy_summary_only_cache_is_not_reused(self) -> None:
        requested_start = date(2026, 7, 1)
        requested_end = date(2026, 7, 2)

        with TemporaryDirectory() as temporary_directory:
            with patch.object(app, "CACHE_DIR", Path(temporary_directory)):
                cache_path = app._cache_path("000001", requested_start, requested_end)
                cache_path.parent.mkdir(parents=True)
                cache_path.write_text(
                    json.dumps(
                        {
                            "version": 2,
                            "source": "legacy",
                            "requested_start": requested_start.isoformat(),
                            "requested_end": requested_end.isoformat(),
                            "change_pct": 10.0,
                        }
                    ),
                    encoding="utf-8",
                )
                cached = app._read_cached_outcome(
                    "000001", requested_start, requested_end, cache_hours=1.0
                )

        self.assertEqual(app.CACHE_VERSION, 3)
        self.assertIsNone(cached)

    def test_zero_cache_hours_does_not_write_a_cache_record(self) -> None:
        history = close_history(
            [("2026-06-30", 10.0), ("2026-07-01", 11.0), ("2026-07-02", 12.0)]
        )
        with (
            patch.object(app, "fetch_period_history", return_value=(history, "test-source")),
            patch.object(app, "_write_cached_outcome") as cache_writer,
        ):
            outcome = app.fetch_period_return(
                "000001",
                date(2026, 7, 1),
                date(2026, 7, 2),
                cache_hours=0.0,
                force_refresh=True,
                limiter=app.RequestRateLimiter(0.0),
                eastmoney_breaker=app.SourceCircuitBreaker("eastmoney"),
                tencent_breaker=app.SourceCircuitBreaker("tencent"),
                timeout_seconds=1.0,
            )

        self.assertAlmostEqual(outcome.change_pct, (12.0 / 11.0 - 1.0) * 100.0)
        self.assertEqual(len(outcome.daily_returns), 2)
        cache_writer.assert_not_called()


class PeriodReturnCollectionAndOutputTests(unittest.TestCase):
    def test_collection_emits_daily_rows_summary_and_failure_rows_in_stable_order(self) -> None:
        sequence_column, code_column, name_column = app.SUMMARY_COLUMNS[:3]
        companies = pd.DataFrame(
            [
                {sequence_column: 2, code_column: "000001", name_column: "Alpha"},
                {sequence_column: 1, code_column: "000003", name_column: "Gamma"},
                {sequence_column: 3, code_column: "000002", name_column: "Broken"},
            ]
        )
        outcomes = {
            "000001": outcome_from_history(
                "000001",
                close_history([("2026-06-30", 10.0), ("2026-07-01", 11.0), ("2026-07-02", 12.0)]),
            ),
            "000003": outcome_from_history(
                "000003",
                close_history([("2026-06-30", 5.0), ("2026-07-01", 5.5), ("2026-07-02", 5.0)]),
            ),
        }

        def fetch(code: str, *args: object, **kwargs: object) -> app.PeriodReturnOutcome:
            if code == "000002":
                raise app.PeriodReturnError("expected failure")
            return outcomes[code]

        with patch.object(app, "fetch_period_return", side_effect=fetch):
            summary_results, daily_results, failures, summary = app.collect_period_returns(
                companies,
                requested_start=date(2026, 7, 1),
                requested_end=date(2026, 7, 2),
                max_companies=None,
                cache_hours=1.0,
                force_refresh=False,
                workers=1,
                request_interval_seconds=0.0,
                timeout_seconds=1.0,
            )

        self.assertEqual(list(summary.values()), [3, 2, 1, 0, 4])
        self.assertEqual(summary_results[code_column].tolist(), ["000001", "000003"])
        self.assertEqual(
            daily_results[code_column].tolist(),
            ["000003", "000001", "000003", "000001"],
        )
        self.assertEqual(daily_results[app.DAILY_RETURN_COLUMNS[3]].tolist(), [
            date(2026, 7, 1),
            date(2026, 7, 1),
            date(2026, 7, 2),
            date(2026, 7, 2),
        ])
        self.assertEqual(failures[code_column].tolist(), ["000002"])

    def test_build_wide_results_keeps_each_trade_day_and_sorts_by_total_return(self) -> None:
        summary_results = pd.DataFrame(
            [
                row(
                    app.SUMMARY_COLUMNS,
                    [
                        2,
                        "000002",
                        "Beta",
                        date(2026, 7, 1),
                        date(2026, 7, 1),
                        10.0,
                        date(2026, 7, 2),
                        date(2026, 7, 2),
                        10.5,
                        5.0,
                        "source-beta",
                        True,
                    ],
                ),
                row(
                    app.SUMMARY_COLUMNS,
                    [
                        3,
                        "000003",
                        "Gamma",
                        date(2026, 7, 1),
                        date(2026, 7, 1),
                        10.0,
                        date(2026, 7, 2),
                        date(2026, 7, 2),
                        12.0,
                        20.0,
                        "source-gamma",
                        False,
                    ],
                ),
                row(
                    app.SUMMARY_COLUMNS,
                    [
                        1,
                        "000001",
                        "Alpha",
                        date(2026, 7, 1),
                        date(2026, 7, 1),
                        10.0,
                        date(2026, 7, 2),
                        date(2026, 7, 2),
                        12.0,
                        20.0,
                        "source-alpha",
                        False,
                    ],
                ),
            ]
        )
        daily_results = pd.DataFrame(
            [
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        2,
                        "000002",
                        "Beta",
                        date(2026, 7, 2),
                        date(2026, 7, 1),
                        10.0,
                        10.5,
                        5.0,
                        "source-beta",
                        True,
                    ],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        3,
                        "000003",
                        "Gamma",
                        date(2026, 7, 1),
                        date(2026, 6, 30),
                        10.0,
                        11.0,
                        10.0,
                        "source-gamma",
                        False,
                    ],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        1,
                        "000001",
                        "Alpha",
                        date(2026, 7, 2),
                        date(2026, 7, 1),
                        11.0,
                        12.0,
                        (12.0 / 11.0 - 1.0) * 100.0,
                        "source-alpha",
                        False,
                    ],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        1,
                        "000001",
                        "Alpha",
                        date(2026, 7, 1),
                        date(2026, 6, 30),
                        10.0,
                        11.0,
                        10.0,
                        "source-alpha",
                        False,
                    ],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        3,
                        "000003",
                        "Gamma",
                        date(2026, 7, 2),
                        date(2026, 7, 1),
                        11.0,
                        12.0,
                        (12.0 / 11.0 - 1.0) * 100.0,
                        "source-gamma",
                        False,
                    ],
                ),
            ]
        )

        wide_results = app.build_wide_results(summary_results, daily_results)

        self.assertEqual(
            wide_results.columns.tolist(),
            [
                "序号",
                "股票代码",
                "股票名称",
                "总涨跌幅（%）",
                "数据来源",
                "缓存命中",
                "2026-07-01",
                "2026-07-02",
            ],
        )
        self.assertEqual(wide_results["股票代码"].tolist(), ["000001", "000003", "000002"])
        self.assertEqual(wide_results["总涨跌幅（%）"].tolist(), [20.0, 20.0, 5.0])
        self.assertEqual(wide_results.loc[0, "2026-07-01"], 10.0)
        self.assertAlmostEqual(wide_results.loc[0, "2026-07-02"], (12.0 / 11.0 - 1.0) * 100.0)
        self.assertEqual(wide_results.loc[1, "2026-07-01"], 10.0)
        self.assertAlmostEqual(wide_results.loc[1, "2026-07-02"], (12.0 / 11.0 - 1.0) * 100.0)
        self.assertTrue(pd.isna(wide_results.loc[2, "2026-07-01"]))
        self.assertEqual(wide_results.loc[2, "2026-07-02"], 5.0)

    def test_write_results_workbook_writes_wide_main_sheet_and_optional_detail_sheets(self) -> None:
        summary_results = pd.DataFrame(
            [
                row(
                    app.SUMMARY_COLUMNS,
                    [
                        2,
                        "000001",
                        "Alpha",
                        date(2026, 7, 1),
                        date(2026, 7, 1),
                        10.0,
                        date(2026, 7, 2),
                        date(2026, 7, 2),
                        12.0,
                        20.0,
                        "test-source",
                        False,
                    ],
                ),
                row(
                    app.SUMMARY_COLUMNS,
                    [
                        1,
                        "000003",
                        "Gamma",
                        date(2026, 7, 1),
                        date(2026, 7, 1),
                        10.0,
                        date(2026, 7, 2),
                        date(2026, 7, 2),
                        10.5,
                        5.0,
                        "other-source",
                        True,
                    ],
                ),
            ]
        )
        daily_results = pd.DataFrame(
            [
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        1,
                        "000001",
                        "Alpha",
                        date(2026, 7, 1),
                        date(2026, 6, 30),
                        10.0,
                        11.0,
                        10.0,
                        "test-source",
                        False,
                    ],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        2,
                        "000001",
                        "Alpha",
                        date(2026, 7, 2),
                        date(2026, 7, 1),
                        11.0,
                        12.0,
                        (12.0 / 11.0 - 1.0) * 100.0,
                        "test-source",
                        False,
                    ],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [
                        1,
                        "000003",
                        "Gamma",
                        date(2026, 7, 2),
                        date(2026, 7, 1),
                        10.0,
                        10.5,
                        5.0,
                        "other-source",
                        True,
                    ],
                ),
            ]
        )
        failures = pd.DataFrame(
            [row(app.FAILURE_COLUMNS, [2, "000002", "Broken", "no history"])]
        )

        with TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "daily_returns.xlsx"
            app.write_results_workbook(summary_results, daily_results, failures, output_path)
            workbook = load_workbook(output_path)
            wide_sheet, detail_sheet, summary_sheet, failure_sheet = workbook.worksheets
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["每日涨跌幅", "每日涨跌幅明细", "区间汇总", "失败明细"],
                )
                self.assertEqual(
                    [cell.value for cell in wide_sheet[1]],
                    [
                        "序号",
                        "股票代码",
                        "股票名称",
                        "总涨跌幅（%）",
                        "数据来源",
                        "缓存命中",
                        "2026-07-01",
                        "2026-07-02",
                    ],
                )
                self.assertEqual(wide_sheet["B2"].value, "000001")
                self.assertEqual(wide_sheet["D2"].value, 20.0)
                self.assertEqual(wide_sheet["G2"].value, 10.0)
                self.assertAlmostEqual(wide_sheet["H2"].value, (12.0 / 11.0 - 1.0) * 100.0)
                self.assertEqual(wide_sheet["B3"].value, "000003")
                self.assertIsNone(wide_sheet["G3"].value)
                self.assertEqual(wide_sheet["H3"].value, 5.0)
                self.assertEqual(wide_sheet["B2"].number_format, "@")
                self.assertEqual(wide_sheet["D2"].number_format, "0.00")
                self.assertEqual(wide_sheet["G2"].number_format, "0.00")
                self.assertEqual(pd.Timestamp(detail_sheet["D2"].value).date(), date(2026, 7, 1))
                self.assertEqual(pd.Timestamp(detail_sheet["E2"].value).date(), date(2026, 6, 30))
                self.assertEqual(detail_sheet["F2"].value, 10.0)
                self.assertEqual(detail_sheet["H2"].value, 10.0)
                self.assertEqual(detail_sheet["D2"].number_format, "yyyy-mm-dd")
                self.assertEqual(detail_sheet["H2"].number_format, "0.00")
                self.assertEqual(summary_sheet["J2"].value, 20.0)
                self.assertEqual(summary_sheet["J2"].number_format, "0.00")
                self.assertEqual(failure_sheet["B2"].value, "000002")
                self.assertEqual(wide_sheet.freeze_panes, "G2")
            finally:
                workbook.close()

    def test_daily_detail_output_splits_before_excel_row_limit(self) -> None:
        daily_results = pd.DataFrame(
            [
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [1, "000001", "Alpha", date(2026, 7, 1), None, None, 10.0, None, "source", False],
                ),
                row(
                    app.DAILY_RETURN_COLUMNS,
                    [2, "000002", "Beta", date(2026, 7, 1), None, None, 20.0, None, "source", False],
                ),
            ]
        )

        with TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "split_daily_returns.xlsx"
            with patch.object(app, "MAX_EXCEL_DATA_ROWS", 1):
                app.write_results_workbook(
                    pd.DataFrame(columns=app.SUMMARY_COLUMNS),
                    daily_results,
                    pd.DataFrame(columns=app.FAILURE_COLUMNS),
                    output_path,
                )
            workbook = load_workbook(output_path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "每日涨跌幅",
                        "每日涨跌幅明细_1",
                        "每日涨跌幅明细_2",
                        "区间汇总",
                        "失败明细",
                    ],
                )
                self.assertEqual(workbook["每日涨跌幅"].max_row, 1)
                self.assertEqual(workbook["每日涨跌幅明细_1"].max_row, 2)
                self.assertEqual(workbook["每日涨跌幅明细_2"].max_row, 2)
            finally:
                workbook.close()

    def test_parse_args_accepts_explicit_period(self) -> None:
        args = app.parse_args(["--start-date", "2026-07-01", "--end-date", "2026-07-23"])

        self.assertEqual(args.start_date, date(2026, 7, 1))
        self.assertEqual(args.end_date, date(2026, 7, 23))


if __name__ == "__main__":
    unittest.main()
