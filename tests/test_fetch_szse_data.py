from __future__ import annotations

import os
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import fetch_szse_data as fetcher


class SessionConfigurationTests(unittest.TestCase):
    def test_build_session_ignores_proxy_configuration(self) -> None:
        with patch.dict(
            os.environ,
            {
                "HTTP_PROXY": "http://127.0.0.1:7890",
                "HTTPS_PROXY": "http://127.0.0.1:7890",
            },
            clear=False,
        ):
            session = fetcher.build_session()
            settings = session.merge_environment_settings(
                fetcher.SZSE_REPORT_URL, {}, None, True, None
            )

        self.assertIs(session.trust_env, False)
        self.assertFalse(settings["proxies"])


class MainboardCompanyIndustryTests(unittest.TestCase):
    def test_mainboard_companies_include_industry_from_sshymc(self) -> None:
        rows = [
            {
                "bk": "主板",
                "agdm": '<a href="/market">000001</a>',
                "agjc": "平安&nbsp;银行",
                "sshymc": "J 金融业",
            },
            {
                "bk": "创业板",
                "agdm": "300001",
                "agjc": "特锐德",
                "sshymc": "制造业",
            },
        ]

        companies = fetcher.build_mainboard_company_dataframe(rows)

        self.assertEqual(companies.columns.tolist(), ["公司代码", "公司简称", "所属行业"])
        self.assertEqual(
            companies.to_dict("records"),
            [
                {
                    "公司代码": "000001",
                    "公司简称": "平安 银行",
                    "所属行业": "J 金融业",
                }
            ],
        )

    def test_missing_or_blank_industry_rejects_stock_pool(self) -> None:
        for industry in (None, " \u00a0 \t"):
            with self.subTest(industry=industry):
                with self.assertRaisesRegex(fetcher.SzseApiError, "sshymc"):
                    fetcher.build_mainboard_company_dataframe(
                        [
                            {
                                "bk": "主板",
                                "agdm": "000002",
                                "agjc": "测试二",
                                "sshymc": industry,
                            }
                        ]
                    )

    def test_mainboard_company_code_must_be_exactly_six_digits(self) -> None:
        with self.assertRaisesRegex(fetcher.SzseApiError, "6 位数字"):
            fetcher.build_mainboard_company_dataframe(
                [
                    {
                        "bk": "主板",
                        "agdm": "000001A",
                        "agjc": "测试公司",
                        "sshymc": "C 制造业",
                    }
                ]
            )

    def test_duplicate_mainboard_company_code_rejects_stock_pool(self) -> None:
        with self.assertRaisesRegex(fetcher.SzseApiError, "代码重复"):
            fetcher.build_mainboard_company_dataframe(
                [
                    {
                        "bk": "主板",
                        "agdm": "000001",
                        "agjc": "测试公司一",
                        "sshymc": "C 制造业",
                    },
                    {
                        "bk": "主板",
                        "agdm": "000001",
                        "agjc": "测试公司二",
                        "sshymc": "C 制造业",
                    },
                ]
            )

    def test_invalid_mainboard_data_fails_before_workbook_write(self) -> None:
        etf_rows = [
            {"sys_key": "159001", "kzjcurl": "测试ETF", "nhzs": "测试指数"}
        ]
        invalid_company_rows = [
            {
                "bk": "主板",
                "agdm": "000001",
                "agjc": "测试公司",
                "sshymc": "",
            }
        ]

        with TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "stock_pool.xlsx"
            with (
                patch.object(fetcher, "build_session"),
                patch.object(
                    fetcher,
                    "fetch_report_rows",
                    side_effect=[etf_rows, invalid_company_rows],
                ),
                patch.object(fetcher, "write_workbook") as write_workbook,
            ):
                with self.assertRaises(fetcher.SzseApiError):
                    fetcher.main(["--output", str(output_path)])

        write_workbook.assert_not_called()

    def test_workbook_writes_company_industry_column(self) -> None:
        company_frame = fetcher.build_mainboard_company_dataframe(
            [
                {
                    "bk": "主板",
                    "agdm": "000001",
                    "agjc": "平安银行",
                    "sshymc": "J 金融业",
                }
            ]
        )
        etf_frame, _, _ = fetcher.build_etf_dataframe(
            [{"sys_key": "159001", "kzjcurl": "测试ETF", "nhzs": "测试指数"}]
        )

        with TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "stock_pool.xlsx"
            fetcher.write_workbook(etf_frame, company_frame, output_path)
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                worksheet = workbook["主板公司"]
                headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
                values = [
                    cell.value
                    for cell in next(worksheet.iter_rows(min_row=2, max_row=2))
                ]
            finally:
                workbook.close()

        self.assertEqual(headers, ["公司代码", "公司简称", "所属行业"])
        self.assertEqual(values, ["000001", "平安银行", "J 金融业"])


if __name__ == "__main__":
    unittest.main()
