"""从深交所下载 ETF 和主板公司列表，并保存到 Excel 工作簿。

在本目录运行：

    uv run python fetch_szse_data.py

默认在脚本所在目录生成 ``深交所数据.xlsx``。数据来自深交所公开 JSON 报表接口，
不使用浏览器自动化或网页 HTML 抓取。
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
import time
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SZSE_REPORT_URL = "https://www.szse.cn/api/report/ShowReport/data"
ETF_CATALOG_ID = "1945"
STOCK_CATALOG_ID = "1110"
TAB_KEY = "tab1"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "深交所数据.xlsx"
DEFAULT_TIMEOUT_SECONDS = 20.0
DEFAULT_REQUEST_INTERVAL_SECONDS = 0.12
DEFAULT_MAX_RETRIES = 3

# 深交所公开接口需要桌面浏览器形式的 User-Agent。
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

# 这些词保持明确且便于扩展，同时匹配基金简称和跟踪指数。命中后即从最终的
# 境内 ETF 表中剔除。
CROSS_BORDER_KEYWORDS = (
    "港股",
    "香港",
    "港股通",
    "沪港深",
    "红筹",
    "恒生",
    "恒指",
    "H股",
    "HANG SENG",
    "HSI",
    "HSCI",
    "HSCEI",
    "HSTECH",
    "美股",
    "美国",
    "纳斯达克",
    "纳指",
    "标普",
    "道琼斯",
    "罗素",
    "中概",
    "日经",
    "日本",
    "东证",
    "印度",
    "越南",
    "泰国",
    "韩国",
    "新加坡",
    "德国",
    "法国",
    "欧洲",
    "欧股",
    "英国",
    "澳洲",
    "亚太",
    "全球",
    "海外",
    "境外",
    "QDII",
    "沙特",
    "巴西",
    "墨西哥",
    "土耳其",
    "美元",
    # 官方 ETF 数据中出现过的明确境外指数标识。不要使用 "HSC"、"MSCI"
    # 等宽泛前缀，因为它们也可能出现在境内指数标识中。
    "HSBIO",
    "HSHCI",
    "HSHKBIO",
    "HSCGSI",
    "HSHYLV",
    "HSIII",
    "HSISC",
    "HSSCAM",
    "HSSCITI",
    "HSSCID",
    "HSSCHI",
    "HSSCSOY",
    "HSMCHYI",
    "SPXNTR",
    "SP5CSSUP",
    "SPSIBI",
    "SPSIOP",
    "NDX",
    "N225",
    "DAX",
    "IBOVESPA",
    "FISAULM",
    "GPCSP006",
)

_TAG_PATTERN = re.compile(r"<[^>]*>")
_SIX_DIGIT_CODE_PATTERN = re.compile(r"(?<!\d)(\d{6})(?!\d)")
_WHITESPACE_PATTERN = re.compile(r"\s+")


class SzseApiError(RuntimeError):
    """深交所公开接口返回的数据无法安全使用时抛出的异常。"""


def build_session() -> requests.Session:
    """创建带有深交所公开 JSON 接口所需请求头的会话。"""

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": "https://www.szse.cn/market/product/list/index.html",
        }
    )
    return session


def fetch_report_rows(
    session: requests.Session,
    *,
    catalog_id: str,
    extra_params: Mapping[str, str],
    timeout_seconds: float,
    request_interval_seconds: float,
    max_retries: int,
) -> list[dict[str, Any]]:
    """获取一个深交所官方报表标签页中的全部分页数据。

    该接口固定每页条数，因此必须根据 ``metadata.pagecount`` 逐页获取，不能
    依赖请求中指定的分页大小。最终记录数与元数据不一致时直接报错，避免生成
    不完整但表面正常的工作簿。
    """

    base_params = {
        "SHOWTYPE": "JSON",
        "CATALOGID": catalog_id,
        "TABKEY": TAB_KEY,
        **dict(extra_params),
    }
    first_tab = _fetch_report_tab(
        session,
        {**base_params, "PAGENO": "1"},
        timeout_seconds=timeout_seconds,
        max_retries=max_retries,
    )
    metadata = _metadata(first_tab)
    page_count = _as_positive_int(metadata.get("pagecount"), "pagecount")
    record_count = _as_non_negative_int(metadata.get("recordcount"), "recordcount")
    rows = _data_rows(first_tab)

    for page_number in range(2, page_count + 1):
        if request_interval_seconds:
            time.sleep(request_interval_seconds)
        page_tab = _fetch_report_tab(
            session,
            {**base_params, "PAGENO": str(page_number)},
            timeout_seconds=timeout_seconds,
            max_retries=max_retries,
        )
        rows.extend(_data_rows(page_tab))

    if len(rows) != record_count:
        raise SzseApiError(
            f"深交所目录 {catalog_id} 返回 {len(rows)} 条记录，"
            f"但元数据声明为 {record_count} 条，请稍后重试。"
        )
    return rows


def _fetch_report_tab(
    session: requests.Session,
    params: Mapping[str, str],
    *,
    timeout_seconds: float,
    max_retries: int,
) -> Mapping[str, Any]:
    payload = _get_json(
        session,
        params,
        timeout_seconds=timeout_seconds,
        max_retries=max_retries,
    )
    if not isinstance(payload, list):
        raise SzseApiError("深交所响应不是预期的报表标签数组。")

    for tab in payload:
        if isinstance(tab, Mapping) and _metadata(tab).get("tabkey") == TAB_KEY:
            return tab
    raise SzseApiError(f"深交所响应中不包含标签 {TAB_KEY!r}。")


def _get_json(
    session: requests.Session,
    params: Mapping[str, str],
    *,
    timeout_seconds: float,
    max_retries: int,
) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            response = session.get(
                SZSE_REPORT_URL,
                params=dict(params),
                timeout=timeout_seconds,
            )
            response.raise_for_status()
            # 深交所声明使用 UTF-8。直接按字节解码，避免网络代理遗漏响应元数据
            # 时中文名称发生乱码。
            return json.loads(response.content.decode("utf-8-sig"))
        except (requests.RequestException, UnicodeDecodeError, json.JSONDecodeError) as exc:
            last_error = exc
            if attempt == max_retries:
                break
            time.sleep(min(2**(attempt - 1), 4))

    raise SzseApiError(
        f"连续 {max_retries} 次均未能从深交所公开接口获取有效 JSON：{last_error}"
    )


def build_etf_dataframe(rows: Sequence[Mapping[str, Any]]) -> tuple[pd.DataFrame, int, int]:
    """构造境内 ETF 表，并返回跨境和无跟踪指数的剔除数量。"""

    domestic_rows: list[dict[str, str]] = []
    excluded_cross_border_count = 0
    excluded_missing_tracking_index_count = 0
    for row in rows:
        code = _extract_six_digit_code(row.get("sys_key"))
        name = _clean_json_display_text(row.get("kzjcurl"))
        tracking_index = _clean_json_display_text(row.get("nhzs"))
        _require_value(code, "ETF code")
        _require_value(name, f"ETF name for {code}")

        if cross_border_matches(name, tracking_index):
            excluded_cross_border_count += 1
            continue
        # 默认不保留未提供跟踪指数的 ETF，例如货币类 ETF。
        if not tracking_index:
            excluded_missing_tracking_index_count += 1
            continue
        domestic_rows.append(
            {
                "基金代码": code,
                "基金简称": name,
                "跟踪指数": tracking_index,
            }
        )

    frame = pd.DataFrame(domestic_rows, columns=["基金代码", "基金简称", "跟踪指数"])
    frame = frame.drop_duplicates(subset=["基金代码"], keep="first")
    return (
        frame.sort_values("基金代码", kind="stable").reset_index(drop=True),
        excluded_cross_border_count,
        excluded_missing_tracking_index_count,
    )


def build_mainboard_company_dataframe(rows: Sequence[Mapping[str, Any]]) -> pd.DataFrame:
    """从 A 股报表构造主板公司表，并防御性过滤非主板记录。"""

    companies: list[dict[str, str]] = []
    for row in rows:
        if _clean_json_display_text(row.get("bk")) != "主板":
            continue
        code = _extract_six_digit_code(row.get("agdm"))
        name = _clean_json_display_text(row.get("agjc"))
        _require_value(code, "main-board company code")
        _require_value(name, f"main-board company name for {code}")
        companies.append({"公司代码": code, "公司简称": name})

    frame = pd.DataFrame(companies, columns=["公司代码", "公司简称"])
    frame = frame.drop_duplicates(subset=["公司代码"], keep="first")
    return frame.sort_values("公司代码", kind="stable").reset_index(drop=True)


def cross_border_matches(fund_name: str, tracking_index: str) -> tuple[str, ...]:
    """返回基金简称和跟踪指数中命中的跨境关键词。"""

    searchable_text = f"{fund_name} {tracking_index}".upper()
    return tuple(
        keyword
        for keyword in CROSS_BORDER_KEYWORDS
        if keyword.upper() in searchable_text
    )


def write_workbook(
    etf_frame: pd.DataFrame,
    company_frame: pd.DataFrame,
    output_path: Path,
) -> None:
    """写入所需的两个 Excel 工作表，并应用基础易读格式。"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        etf_frame.to_excel(writer, sheet_name="ETF", index=False)
        company_frame.to_excel(writer, sheet_name="主板公司", index=False)

        for sheet_name in ("ETF", "主板公司"):
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            _format_worksheet(worksheet)


def _format_worksheet(worksheet: Any) -> None:
    """让生成工作簿的表头可见，并保持代码列为文本格式。"""

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[column_letter].width = min(max(width, 12), 36)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for cell in worksheet["A"][1:]:
        cell.number_format = "@"


def _metadata(tab: Mapping[str, Any]) -> Mapping[str, Any]:
    metadata = tab.get("metadata")
    if not isinstance(metadata, Mapping):
        raise SzseApiError("深交所报表标签页缺少 metadata 对象。")
    return metadata


def _data_rows(tab: Mapping[str, Any]) -> list[dict[str, Any]]:
    data = tab.get("data")
    if not isinstance(data, list):
        raise SzseApiError("深交所报表标签页缺少 data 数组。")
    if not all(isinstance(row, Mapping) for row in data):
        raise SzseApiError("深交所报表数据中包含非对象记录。")
    return [dict(row) for row in data]


def _as_positive_int(value: Any, field_name: str) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise SzseApiError(f"深交所元数据 {field_name!r} 无效：{value!r}") from exc
    if number < 1:
        raise SzseApiError(f"深交所元数据 {field_name!r} 必须为正数：{value!r}")
    return number


def _as_non_negative_int(value: Any, field_name: str) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise SzseApiError(f"深交所元数据 {field_name!r} 无效：{value!r}") from exc
    if number < 0:
        raise SzseApiError(f"深交所元数据 {field_name!r} 不能为负数：{value!r}")
    return number


def _clean_json_display_text(value: Any) -> str:
    """清除 JSON 字段中用于展示的标记，并非解析网页 HTML。

    官方 JSON 会在少数表格字段中附加 ``<a>``、``<u>`` 标签。本函数只规范化
    这些 JSON 字符串；脚本不会下载、跟随或解析任何 HTML 文档。
    """

    if value is None:
        return ""
    text = html.unescape(str(value))
    text = _TAG_PATTERN.sub("", text).replace("\xa0", " ")
    return _WHITESPACE_PATTERN.sub(" ", text).strip()


def _extract_six_digit_code(value: Any) -> str:
    match = _SIX_DIGIT_CODE_PATTERN.search(_clean_json_display_text(value))
    return match.group(1) if match else ""


def _require_value(value: str, field_name: str) -> None:
    if not value:
        raise SzseApiError(f"深交所响应缺少 {field_name}。")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="获取深交所境内 ETF 和主板公司，并导出到 Excel（默认剔除无跟踪指数 ETF）。"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Excel 输出路径（默认：深交所数据.xlsx）。",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_TIMEOUT_SECONDS,
        help="单次请求超时秒数（默认：20）。",
    )
    parser.add_argument(
        "--interval",
        type=float,
        default=DEFAULT_REQUEST_INTERVAL_SECONDS,
        help="报表分页请求之间的间隔秒数（默认：0.12）。",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.timeout <= 0:
        raise ValueError("--timeout 必须为正数。")
    if args.interval < 0:
        raise ValueError("--interval 不能为负数。")

    session = build_session()
    print("正在获取深交所 ETF 列表...")
    etf_rows = fetch_report_rows(
        session,
        catalog_id=ETF_CATALOG_ID,
        extra_params={},
        timeout_seconds=args.timeout,
        request_interval_seconds=args.interval,
        max_retries=DEFAULT_MAX_RETRIES,
    )
    (
        etf_frame,
        excluded_cross_border_count,
        excluded_missing_tracking_index_count,
    ) = build_etf_dataframe(etf_rows)

    print("正在获取深交所主板公司列表...")
    company_rows = fetch_report_rows(
        session,
        catalog_id=STOCK_CATALOG_ID,
        extra_params={"selectModule": "main"},
        timeout_seconds=args.timeout,
        request_interval_seconds=args.interval,
        max_retries=DEFAULT_MAX_RETRIES,
    )
    company_frame = build_mainboard_company_dataframe(company_rows)
    if company_frame.empty:
        raise SzseApiError("深交所主板筛选结果为空。")

    output_path = args.output.resolve()
    write_workbook(etf_frame, company_frame, output_path)
    print(f"已保存：{output_path}")
    print(
        "统计："
        f"ETF 源数据={len(etf_rows)}，跨境剔除={excluded_cross_border_count}，"
        f"无跟踪指数剔除={excluded_missing_tracking_index_count}，"
        f"保留境内 ETF={len(etf_frame)}，主板公司={len(company_frame)}"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (SzseApiError, ValueError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1) from exc
