"""导出深市主板股票在指定区间内每个交易日的前复权涨跌幅。

运行示例：

    uv run --locked python fetch_period_returns.py \
        --start-date 2026-07-01 --end-date 2026-07-23

结果只包含请求区间内的交易日。单日涨跌幅按连续前复权收盘价计算：
((当日收盘价 / 前一交易日收盘价) - 1) * 100。
为了计算区间内第一个交易日，脚本会额外请求开始日前的一段日线数据。
"""

from __future__ import annotations

import argparse
import json
import math
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MODULE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK_PATH = MODULE_DIR / "深交所数据.xlsx"
CACHE_DIR = MODULE_DIR / "data_cache" / "period_returns"

EASTMONEY_DAILY_KLINE_URL = "https://push2his.eastmoney.com/api/qt/stock/kline/get"
TENCENT_RICH_DAILY_KLINE_URLS = (
    "https://proxy.finance.qq.com/ifzqgtimg/appstock/app/newfqkline/get",
    "https://web.ifzq.gtimg.cn/appstock/app/newfqkline/get",
)
TENCENT_LEGACY_DAILY_KLINE_URL = "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"
EASTMONEY_FIELDS1 = "f1,f2,f3,f4,f5,f6"
EASTMONEY_FIELDS2 = "f51,f52,f53"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

DEFAULT_CACHE_HOURS = 12.0
DEFAULT_REQUEST_INTERVAL_SECONDS = 0.25
DEFAULT_WORKERS = 4
MAX_WORKERS = 6
MAX_RETRIES = 3
SOURCE_FAILURE_THRESHOLD = 3
SOURCE_COOLDOWN_SECONDS = 30.0
CACHE_VERSION = 3

# 腾讯接口最多稳定返回约 640 根日线。600 个自然日通常不会超过该限制。
MAX_SINGLE_QUERY_CALENDAR_DAYS = 600
# 相邻分页重叠一段日期，避免接口边界截断导致交易日缺失。
QUERY_WINDOW_OVERLAP_DAYS = 30
# 在请求起点前保留足够的日历日，用来获取首个交易日的前收盘价。
HISTORY_LOOKBACK_CALENDAR_DAYS = 31
MAX_EXCEL_DATA_ROWS = 1_048_575

DAILY_RETURN_COLUMNS = (
    "序号",
    "股票代码",
    "股票名称",
    "交易日期",
    "前一交易日",
    "前收盘价（前复权）",
    "收盘价（前复权）",
    "当日涨跌幅（%）",
    "数据来源",
    "缓存命中",
)
WIDE_RETURN_FIXED_COLUMNS = (
    "序号",
    "股票代码",
    "股票名称",
    "总涨跌幅（%）",
    "数据来源",
    "缓存命中",
)
SUMMARY_COLUMNS = (
    "序号",
    "股票代码",
    "股票名称",
    "请求开始日期",
    "实际开始交易日",
    "开始收盘价（前复权）",
    "请求结束日期",
    "实际结束交易日",
    "结束收盘价（前复权）",
    "区间涨跌幅（%）",
    "数据来源",
    "缓存命中",
)
FAILURE_COLUMNS = ("序号", "股票代码", "股票名称", "失败原因")


class PeriodReturnError(RuntimeError):
    """表示输入或行情数据不足以计算可靠结果。"""


class FetchFailure(PeriodReturnError):
    """记录数据源失败是否应触发共享熔断器。"""

    def __init__(self, message: str, *, service_unavailable: bool) -> None:
        super().__init__(message)
        self.service_unavailable = service_unavailable


@dataclass(frozen=True)
class DailyBar:
    """一根已校验的前复权日线。"""

    trade_date: date
    close: float


@dataclass(frozen=True)
class DailyReturnRecord:
    """一个交易日相对于前一实际交易日的涨跌幅。"""

    trade_date: date
    previous_trade_date: date | None
    previous_close: float | None
    close: float
    change_pct: float | None


@dataclass(frozen=True)
class PeriodReturnOutcome:
    """单只股票的区间汇总、每日结果及可持久化日线。"""

    code: str
    source: str
    from_cache: bool
    requested_start: date
    actual_start: date
    start_close: float
    requested_end: date
    actual_end: date
    end_close: float
    change_pct: float
    daily_returns: tuple[DailyReturnRecord, ...] = ()
    bars: tuple[DailyBar, ...] = ()


class RequestRateLimiter:
    """在所有工作线程之间维持统一的最小请求间隔。"""

    def __init__(self, minimum_interval_seconds: float) -> None:
        self.minimum_interval_seconds = max(0.0, float(minimum_interval_seconds))
        self._lock = threading.Lock()
        self._next_allowed_at = 0.0

    def wait(self) -> None:
        with self._lock:
            now = time.monotonic()
            scheduled_at = max(now, self._next_allowed_at)
            self._next_allowed_at = scheduled_at + self.minimum_interval_seconds
        delay = scheduled_at - now
        if delay > 0:
            time.sleep(delay)

    def penalize(self, seconds: float) -> None:
        with self._lock:
            self._next_allowed_at = max(
                self._next_allowed_at,
                time.monotonic() + max(0.0, float(seconds)),
            )


class SourceCircuitBreaker:
    """当数据源持续不可用时，避免重复等待长时间超时。"""

    def __init__(
        self,
        source_name: str,
        *,
        failure_threshold: int = SOURCE_FAILURE_THRESHOLD,
        cooldown_seconds: float = SOURCE_COOLDOWN_SECONDS,
    ) -> None:
        self.source_name = source_name
        self.failure_threshold = max(1, int(failure_threshold))
        self.cooldown_seconds = max(0.0, float(cooldown_seconds))
        self._lock = threading.Lock()
        self._consecutive_failures = 0
        self._open_until = 0.0

    def allow_request(self) -> bool:
        with self._lock:
            return time.monotonic() >= self._open_until

    def record_success(self) -> None:
        with self._lock:
            self._consecutive_failures = 0
            self._open_until = 0.0

    def record_service_failure(self) -> None:
        with self._lock:
            self._consecutive_failures += 1
            if self._consecutive_failures >= self.failure_threshold:
                self._open_until = max(
                    self._open_until,
                    time.monotonic() + self.cooldown_seconds,
                )

    def unavailable_error(self) -> FetchFailure:
        return FetchFailure(
            f"{self.source_name}服务暂时不可用，已跳过本次请求。",
            service_unavailable=True,
        )


_THREAD_LOCAL = threading.local()


def _request_session() -> requests.Session:
    """为每个工作线程复用独立的 HTTP 会话。"""

    session = getattr(_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        session.trust_env = False
        session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Referer": "https://quote.eastmoney.com/",
            }
        )
        _THREAD_LOCAL.session = session
    return session


def _coerce_number(value: object) -> float | None:
    """将接口中的数值转换为有限浮点数。"""

    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip()
    if not text or text in {"-", "--", "None", "null"}:
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _as_six_digit_code(value: object) -> str | None:
    """标准化为六位证券代码。"""

    text = str(value).strip()
    if text.isdigit() and len(text) <= 6:
        return text.zfill(6)
    return None


def parse_iso_date(value: str) -> date:
    """解析命令行中的 ISO 日期。"""

    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("日期必须为 YYYY-MM-DD 格式。") from exc


def load_mainboard_companies(workbook_path: Path) -> pd.DataFrame:
    """读取主板公司股票池，且不修改输入工作簿。"""

    if not workbook_path.is_file():
        raise PeriodReturnError(f"未找到股票池文件：{workbook_path}")
    try:
        frame = pd.read_excel(workbook_path, sheet_name="主板公司", dtype=str)
    except (OSError, ValueError) as exc:
        raise PeriodReturnError(f"无法读取“主板公司”工作表：{exc}") from exc

    required_columns = {"公司代码", "公司简称"}
    missing_columns = required_columns.difference(frame.columns)
    if missing_columns:
        missing = "、".join(sorted(missing_columns))
        raise PeriodReturnError(f"“主板公司”工作表缺少必要列：{missing}")

    companies = pd.DataFrame(
        {
            "股票代码": frame["公司代码"].map(_as_six_digit_code),
            "股票名称": frame["公司简称"].fillna("").astype(str).str.strip(),
        }
    )
    companies = companies.dropna(subset=["股票代码"])
    companies = companies.loc[companies["股票名称"].ne("")]
    companies = companies.drop_duplicates(subset=["股票代码"], keep="first").reset_index(drop=True)
    if companies.empty:
        raise PeriodReturnError("“主板公司”工作表中没有可用股票代码。")
    companies.insert(0, "序号", range(1, len(companies) + 1))
    return companies


def _normalize_close_history(rows: Sequence[Mapping[str, object]]) -> pd.DataFrame:
    """清洗、按日期排序并去重日线收盘价。"""

    frame = pd.DataFrame(rows, columns=["date", "close"])
    if frame.empty:
        raise PeriodReturnError("接口未返回可用日线。")
    frame["date"] = pd.to_datetime(frame["date"], errors="coerce")
    frame["close"] = pd.to_numeric(frame["close"], errors="coerce")
    frame = frame.dropna(subset=["date", "close"])
    frame = frame.loc[frame["close"].gt(0)]
    frame = frame.sort_values("date", kind="stable")
    frame = frame.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
    if frame.empty:
        raise PeriodReturnError("接口日线没有有效收盘价。")
    return frame


def parse_eastmoney_history(payload: Mapping[str, Any]) -> pd.DataFrame:
    """解析东方财富前复权日 K 接口的日期和收盘价。"""

    data = payload.get("data")
    if not isinstance(data, Mapping):
        raise PeriodReturnError("东方财富响应缺少 data 对象。")
    raw_klines = data.get("klines")
    if not isinstance(raw_klines, list):
        raise PeriodReturnError("东方财富响应缺少日 K 列表。")

    rows: list[dict[str, object]] = []
    for raw_kline in raw_klines:
        if not isinstance(raw_kline, str):
            continue
        fields = raw_kline.split(",")
        if len(fields) < 3:
            continue
        rows.append({"date": fields[0], "close": _coerce_number(fields[2])})
    return _normalize_close_history(rows)


def _parse_tencent_payload(payload: Mapping[str, Any], symbol: str) -> pd.DataFrame:
    """解析腾讯接口中指定证券的前复权日线。"""

    if payload.get("code") not in (None, 0, "0"):
        raise PeriodReturnError(f"腾讯接口返回 code={payload.get('code')!r}。")
    data = payload.get("data")
    if not isinstance(data, Mapping):
        raise PeriodReturnError("腾讯响应缺少 data 对象。")
    security_data = data.get(symbol)
    if not isinstance(security_data, Mapping):
        raise PeriodReturnError(f"腾讯响应缺少 {symbol} 的日 K 数据。")
    raw_klines = security_data.get("qfqday")
    if not isinstance(raw_klines, list):
        raise PeriodReturnError("腾讯响应缺少前复权日 K 列表。")

    rows: list[dict[str, object]] = []
    for raw_kline in raw_klines:
        if not isinstance(raw_kline, (list, tuple)) or len(raw_kline) < 3:
            continue
        rows.append({"date": raw_kline[0], "close": _coerce_number(raw_kline[2])})
    return _normalize_close_history(rows)


def parse_tencent_rich_history(raw_text: str, symbol: str) -> pd.DataFrame:
    """解析腾讯增强接口包裹在变量赋值中的 JSON 响应。"""

    json_text = raw_text.strip()
    if "=" in json_text:
        json_text = json_text.split("=", 1)[1].strip()
    json_text = json_text.rstrip(";").strip()
    try:
        payload = json.loads(json_text)
    except json.JSONDecodeError as exc:
        raise PeriodReturnError("腾讯增强日 K 响应不是有效 JSON。") from exc
    if not isinstance(payload, Mapping):
        raise PeriodReturnError("腾讯增强日 K 响应不是对象。")
    return _parse_tencent_payload(payload, symbol)


def select_period_return(
    history: pd.DataFrame,
    requested_start: date,
    requested_end: date,
) -> tuple[date, float, date, float, float]:
    """从请求区间内首尾两个实际交易日计算区间涨跌幅。"""

    if requested_start > requested_end:
        raise PeriodReturnError("开始日期不能晚于结束日期。")
    period_rows = history.loc[
        history["date"].between(pd.Timestamp(requested_start), pd.Timestamp(requested_end))
    ]
    if period_rows.empty:
        raise PeriodReturnError("指定区间内没有可用交易日。")

    start_row = period_rows.iloc[0]
    end_row = period_rows.iloc[-1]
    actual_start = pd.Timestamp(start_row["date"]).date()
    actual_end = pd.Timestamp(end_row["date"]).date()
    start_close = float(start_row["close"])
    end_close = float(end_row["close"])
    if actual_start > actual_end:
        raise PeriodReturnError("实际开始交易日晚于实际结束交易日。")
    if start_close <= 0 or end_close <= 0:
        raise PeriodReturnError("区间边界收盘价无效。")
    change_pct = (end_close / start_close - 1.0) * 100.0
    return actual_start, start_close, actual_end, end_close, change_pct


def select_daily_returns(
    history: pd.DataFrame,
    requested_start: date,
    requested_end: date,
) -> tuple[DailyReturnRecord, ...]:
    """计算请求区间内每个交易日相对前一交易日的前复权涨跌幅。"""

    if requested_start > requested_end:
        raise PeriodReturnError("开始日期不能晚于结束日期。")

    calculated = history.copy()
    calculated = calculated.sort_values("date", kind="stable")
    calculated = calculated.drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
    calculated["previous_date"] = calculated["date"].shift(1)
    calculated["previous_close"] = calculated["close"].shift(1)
    calculated["change_pct"] = calculated["close"].pct_change() * 100.0
    period_rows = calculated.loc[
        calculated["date"].between(pd.Timestamp(requested_start), pd.Timestamp(requested_end))
    ]
    if period_rows.empty:
        raise PeriodReturnError("指定区间内没有可用交易日。")

    records: list[DailyReturnRecord] = []
    for row in period_rows.itertuples(index=False):
        previous_date = None if pd.isna(row.previous_date) else pd.Timestamp(row.previous_date).date()
        previous_close = None if pd.isna(row.previous_close) else float(row.previous_close)
        change_pct = None if pd.isna(row.change_pct) else float(row.change_pct)
        records.append(
            DailyReturnRecord(
                trade_date=pd.Timestamp(row.date).date(),
                previous_trade_date=previous_date,
                previous_close=previous_close,
                close=float(row.close),
                change_pct=change_pct,
            )
        )
    return tuple(records)


def _history_to_bars(history: pd.DataFrame) -> tuple[DailyBar, ...]:
    """将数据框转换为可写入缓存的日线记录。"""

    return tuple(
        DailyBar(trade_date=pd.Timestamp(row.date).date(), close=float(row.close))
        for row in history.itertuples(index=False)
    )


def _bars_to_history(raw_bars: object) -> pd.DataFrame:
    """从缓存中的日线数组恢复已清洗的收盘价序列。"""

    if not isinstance(raw_bars, list):
        raise PeriodReturnError("缓存缺少完整日线数据。")
    rows: list[dict[str, object]] = []
    for raw_bar in raw_bars:
        if not isinstance(raw_bar, Mapping):
            continue
        rows.append({"date": raw_bar.get("date"), "close": raw_bar.get("close")})
    return _normalize_close_history(rows)


def _outcome_from_history(
    code: str,
    source: str,
    from_cache: bool,
    requested_start: date,
    requested_end: date,
    history: pd.DataFrame,
) -> PeriodReturnOutcome:
    """由完整日线统一构造汇总和逐日结果，避免两种结果口径不一致。"""

    actual_start, start_close, actual_end, end_close, change_pct = select_period_return(
        history,
        requested_start,
        requested_end,
    )
    return PeriodReturnOutcome(
        code=code,
        source=source,
        from_cache=from_cache,
        requested_start=requested_start,
        actual_start=actual_start,
        start_close=start_close,
        requested_end=requested_end,
        actual_end=actual_end,
        end_close=end_close,
        change_pct=change_pct,
        daily_returns=select_daily_returns(history, requested_start, requested_end),
        bars=_history_to_bars(history),
    )


def _cache_path(code: str, requested_start: date, requested_end: date) -> Path:
    """返回单只股票、单个请求区间的缓存文件位置。"""

    return CACHE_DIR / f"{requested_start.isoformat()}_{requested_end.isoformat()}" / f"{code}.json"


def _read_cached_outcome(
    code: str,
    requested_start: date,
    requested_end: date,
    cache_hours: float,
) -> PeriodReturnOutcome | None:
    """读取尚未过期的完整日线缓存，并重新生成结果。"""

    target = _cache_path(code, requested_start, requested_end)
    if not target.is_file() or cache_hours <= 0:
        return None
    if time.time() - target.stat().st_mtime > cache_hours * 3600:
        return None
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
        if (
            not isinstance(payload, Mapping)
            or payload.get("version") != CACHE_VERSION
            or payload.get("requested_start") != requested_start.isoformat()
            or payload.get("requested_end") != requested_end.isoformat()
        ):
            return None
        source = payload.get("source")
        if not isinstance(source, str) or not source:
            return None
        history = _bars_to_history(payload.get("bars"))
        return _outcome_from_history(
            code,
            source,
            True,
            requested_start,
            requested_end,
            history,
        )
    except (OSError, TypeError, ValueError, json.JSONDecodeError, PeriodReturnError):
        return None


def _write_cached_outcome(outcome: PeriodReturnOutcome) -> None:
    """原子写入完整前复权日线缓存，供相同区间重复使用。"""

    if not outcome.bars:
        raise PeriodReturnError("无法缓存缺少完整日线的结果。")
    target = _cache_path(outcome.code, outcome.requested_start, outcome.requested_end)
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": CACHE_VERSION,
        "saved_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source": outcome.source,
        "requested_start": outcome.requested_start.isoformat(),
        "requested_end": outcome.requested_end.isoformat(),
        "bars": [
            {"date": bar.trade_date.isoformat(), "close": bar.close}
            for bar in outcome.bars
        ],
    }
    temporary = target.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    temporary.replace(target)


def _is_service_failure(exc: requests.RequestException) -> bool:
    """判断请求异常是否说明服务端或网络暂不可用。"""

    response = getattr(exc, "response", None)
    return response is None or response.status_code in {403, 408, 429} or response.status_code >= 500


def _is_rate_limited(exc: requests.RequestException) -> bool:
    """判断服务端是否要求降低请求频率。"""

    response = getattr(exc, "response", None)
    return response is not None and response.status_code in {403, 429}


def _window_bar_limit(start_day: date, end_day: date) -> int:
    """按窗口长度计算接口需要返回的最大日线根数。"""

    return max(120, min(640, (end_day - start_day).days + 40))


def _query_windows(requested_start: date, requested_end: date) -> tuple[tuple[date, date], ...]:
    """将长区间连续分页，每页均可完整保留其覆盖日期的日线。"""

    if requested_start > requested_end:
        raise PeriodReturnError("开始日期不能晚于结束日期。")

    windows: list[tuple[date, date]] = []
    start_day = requested_start
    while start_day <= requested_end:
        end_day = min(
            requested_end,
            start_day + timedelta(days=MAX_SINGLE_QUERY_CALENDAR_DAYS - 1),
        )
        windows.append((start_day, end_day))
        if end_day == requested_end:
            break
        start_day = end_day - timedelta(days=QUERY_WINDOW_OVERLAP_DAYS - 1)
    return tuple(windows)


def _fetch_eastmoney_window(
    code: str,
    start_day: date,
    end_day: date,
    limiter: RequestRateLimiter,
    timeout_seconds: float,
) -> pd.DataFrame:
    """请求东方财富的一段前复权日线。"""

    parameters = {
        "secid": f"0.{code}",
        "klt": "101",
        "fqt": "1",
        "beg": start_day.strftime("%Y%m%d"),
        "end": end_day.strftime("%Y%m%d"),
        "lmt": str(_window_bar_limit(start_day, end_day)),
        "fields1": EASTMONEY_FIELDS1,
        "fields2": EASTMONEY_FIELDS2,
    }
    problems: list[str] = []
    service_failures = 0
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            limiter.wait()
            response = _request_session().get(
                EASTMONEY_DAILY_KLINE_URL,
                params=parameters,
                timeout=timeout_seconds,
            )
            response.raise_for_status()
            return parse_eastmoney_history(response.json())
        except requests.RequestException as exc:
            if _is_service_failure(exc):
                service_failures += 1
            if _is_rate_limited(exc):
                limiter.penalize(1.5 * attempt)
            problems.append(f"第 {attempt} 次：{exc}")
        except (TypeError, ValueError, PeriodReturnError) as exc:
            problems.append(f"第 {attempt} 次：{exc}")
        if attempt < MAX_RETRIES:
            time.sleep(0.8 * attempt)
    raise FetchFailure("；".join(problems), service_unavailable=service_failures == MAX_RETRIES)


def _fetch_tencent_rich_window(
    code: str,
    start_day: date,
    end_day: date,
    limiter: RequestRateLimiter,
    timeout_seconds: float,
) -> pd.DataFrame:
    """请求腾讯增强接口的一段前复权日线。"""

    symbol = f"sz{code}"
    parameters = {
        "_var": "kline_dayqfq",
        "param": (
            f"{symbol},day,{start_day.isoformat()},{end_day.isoformat()},"
            f"{_window_bar_limit(start_day, end_day)},qfq"
        ),
        "r": f"{time.time():.6f}",
    }
    problems: list[str] = []
    service_failures = 0
    total_requests = 0
    for attempt in range(1, MAX_RETRIES + 1):
        for endpoint in TENCENT_RICH_DAILY_KLINE_URLS:
            total_requests += 1
            try:
                limiter.wait()
                response = _request_session().get(endpoint, params=parameters, timeout=timeout_seconds)
                response.raise_for_status()
                return parse_tencent_rich_history(response.text, symbol)
            except requests.RequestException as exc:
                if _is_service_failure(exc):
                    service_failures += 1
                if _is_rate_limited(exc):
                    limiter.penalize(1.5 * attempt)
                problems.append(f"第 {attempt} 次 {endpoint}：{exc}")
            except (TypeError, ValueError, PeriodReturnError) as exc:
                problems.append(f"第 {attempt} 次 {endpoint}：{exc}")
        if attempt < MAX_RETRIES:
            time.sleep(0.8 * attempt)
    raise FetchFailure(
        "；".join(problems),
        service_unavailable=service_failures == total_requests,
    )


def _fetch_tencent_legacy_window(
    code: str,
    start_day: date,
    end_day: date,
    limiter: RequestRateLimiter,
    timeout_seconds: float,
) -> pd.DataFrame:
    """请求腾讯旧接口的一段前复权日线。"""

    symbol = f"sz{code}"
    parameters = {
        "param": (
            f"{symbol},day,{start_day.isoformat()},{end_day.isoformat()},"
            f"{_window_bar_limit(start_day, end_day)},qfq"
        )
    }
    problems: list[str] = []
    service_failures = 0
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            limiter.wait()
            response = _request_session().get(
                TENCENT_LEGACY_DAILY_KLINE_URL,
                params=parameters,
                timeout=timeout_seconds,
            )
            response.raise_for_status()
            payload = response.json()
            if not isinstance(payload, Mapping):
                raise PeriodReturnError("腾讯旧版日 K 响应不是对象。")
            return _parse_tencent_payload(payload, symbol)
        except requests.RequestException as exc:
            if _is_service_failure(exc):
                service_failures += 1
            if _is_rate_limited(exc):
                limiter.penalize(1.5 * attempt)
            problems.append(f"第 {attempt} 次：{exc}")
        except (TypeError, ValueError, PeriodReturnError) as exc:
            problems.append(f"第 {attempt} 次：{exc}")
        if attempt < MAX_RETRIES:
            time.sleep(0.8 * attempt)
    raise FetchFailure("；".join(problems), service_unavailable=service_failures == MAX_RETRIES)


def _merge_histories(histories: Sequence[pd.DataFrame]) -> pd.DataFrame:
    """合并分页结果，并以日期去重后得到连续日线。"""

    if not histories:
        raise PeriodReturnError("没有可合并的日线数据。")
    records = pd.concat(histories, ignore_index=True).to_dict("records")
    return _normalize_close_history(records)


def _fetch_windows(
    fetch_window: Any,
    code: str,
    windows: Sequence[tuple[date, date]],
    limiter: RequestRateLimiter,
    timeout_seconds: float,
) -> pd.DataFrame:
    """按顺序拉取同一数据源的所有分页，避免混合复权序列。"""

    histories = [
        fetch_window(code, start_day, end_day, limiter, timeout_seconds)
        for start_day, end_day in windows
    ]
    return _merge_histories(histories)


def _fetch_tencent_history(
    code: str,
    windows: Sequence[tuple[date, date]],
    limiter: RequestRateLimiter,
    timeout_seconds: float,
) -> tuple[pd.DataFrame, str]:
    """优先使用腾讯增强接口，失败后回退腾讯旧接口。"""

    try:
        history = _fetch_windows(
            _fetch_tencent_rich_window,
            code,
            windows,
            limiter,
            timeout_seconds,
        )
        return history, "腾讯增强前复权日 K"
    except FetchFailure as rich_error:
        try:
            history = _fetch_windows(
                _fetch_tencent_legacy_window,
                code,
                windows,
                limiter,
                timeout_seconds,
            )
            return history, "腾讯前复权日 K"
        except FetchFailure as legacy_error:
            raise FetchFailure(
                f"腾讯增强日 K 失败：{rich_error}；腾讯旧版日 K 失败：{legacy_error}",
                service_unavailable=(
                    rich_error.service_unavailable and legacy_error.service_unavailable
                ),
            ) from legacy_error


def fetch_period_history(
    code: str,
    requested_start: date,
    requested_end: date,
    limiter: RequestRateLimiter,
    eastmoney_breaker: SourceCircuitBreaker,
    tencent_breaker: SourceCircuitBreaker,
    timeout_seconds: float,
) -> tuple[pd.DataFrame, str]:
    """抓取完整区间和首日计算所需前置日线的前复权行情。"""

    query_start = requested_start - timedelta(days=HISTORY_LOOKBACK_CALENDAR_DAYS)
    windows = _query_windows(query_start, requested_end)
    try:
        if not tencent_breaker.allow_request():
            raise tencent_breaker.unavailable_error()
        history, source = _fetch_tencent_history(code, windows, limiter, timeout_seconds)
        tencent_breaker.record_success()
        return history, source
    except FetchFailure as exc:
        tencent_error = exc
        if tencent_error.service_unavailable:
            tencent_breaker.record_service_failure()

    try:
        if not eastmoney_breaker.allow_request():
            raise eastmoney_breaker.unavailable_error()
        history = _fetch_windows(
            _fetch_eastmoney_window,
            code,
            windows,
            limiter,
            timeout_seconds,
        )
        eastmoney_breaker.record_success()
        return history, "东方财富前复权日 K"
    except FetchFailure as eastmoney_error:
        if eastmoney_error.service_unavailable:
            eastmoney_breaker.record_service_failure()
        raise FetchFailure(
            f"腾讯失败：{tencent_error}；东方财富回退失败：{eastmoney_error}",
            service_unavailable=(
                tencent_error.service_unavailable and eastmoney_error.service_unavailable
            ),
        ) from eastmoney_error


def fetch_period_return(
    code: str,
    requested_start: date,
    requested_end: date,
    *,
    cache_hours: float,
    force_refresh: bool,
    limiter: RequestRateLimiter,
    eastmoney_breaker: SourceCircuitBreaker,
    tencent_breaker: SourceCircuitBreaker,
    timeout_seconds: float,
) -> PeriodReturnOutcome:
    """获取单只股票的逐日结果；相同区间优先使用完整日线缓存。"""

    if not force_refresh:
        cached = _read_cached_outcome(code, requested_start, requested_end, cache_hours)
        if cached is not None:
            return cached

    history, source = fetch_period_history(
        code,
        requested_start,
        requested_end,
        limiter,
        eastmoney_breaker,
        tencent_breaker,
        timeout_seconds,
    )
    outcome = _outcome_from_history(
        code,
        source,
        False,
        requested_start,
        requested_end,
        history,
    )
    if cache_hours > 0:
        _write_cached_outcome(outcome)
    return outcome


def collect_period_returns(
    companies: pd.DataFrame,
    *,
    requested_start: date,
    requested_end: date,
    max_companies: int | None,
    cache_hours: float,
    force_refresh: bool,
    workers: int,
    request_interval_seconds: float,
    timeout_seconds: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, int]]:
    """并发处理股票池，并汇总逐日数据、区间汇总和失败明细。"""

    if requested_start > requested_end:
        raise ValueError("--start-date 不能晚于 --end-date。")
    if max_companies is not None and max_companies <= 0:
        raise ValueError("--max-companies 必须为正数。")

    selected_companies = companies if max_companies is None else companies.head(max_companies)
    records = selected_companies.to_dict("records")
    limiter = RequestRateLimiter(request_interval_seconds)
    eastmoney_breaker = SourceCircuitBreaker("东方财富")
    tencent_breaker = SourceCircuitBreaker("腾讯")
    summary_rows: list[dict[str, object]] = []
    daily_rows: list[dict[str, object]] = []
    failure_rows: list[dict[str, object]] = []
    cache_hits = 0
    completed = 0

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(
                fetch_period_return,
                str(company["股票代码"]),
                requested_start,
                requested_end,
                cache_hours=cache_hours,
                force_refresh=force_refresh,
                limiter=limiter,
                eastmoney_breaker=eastmoney_breaker,
                tencent_breaker=tencent_breaker,
                timeout_seconds=timeout_seconds,
            ): company
            for company in records
        }
        for future in as_completed(futures):
            company = futures[future]
            completed += 1
            try:
                outcome = future.result()
                cache_hits += int(outcome.from_cache)
                summary_rows.append(
                    {
                        "序号": company["序号"],
                        "股票代码": outcome.code,
                        "股票名称": company["股票名称"],
                        "请求开始日期": outcome.requested_start,
                        "实际开始交易日": outcome.actual_start,
                        "开始收盘价（前复权）": outcome.start_close,
                        "请求结束日期": outcome.requested_end,
                        "实际结束交易日": outcome.actual_end,
                        "结束收盘价（前复权）": outcome.end_close,
                        "区间涨跌幅（%）": outcome.change_pct,
                        "数据来源": outcome.source,
                        "缓存命中": outcome.from_cache,
                    }
                )
                daily_rows.extend(
                    {
                        "序号": company["序号"],
                        "股票代码": outcome.code,
                        "股票名称": company["股票名称"],
                        "交易日期": daily_return.trade_date,
                        "前一交易日": daily_return.previous_trade_date,
                        "前收盘价（前复权）": daily_return.previous_close,
                        "收盘价（前复权）": daily_return.close,
                        "当日涨跌幅（%）": daily_return.change_pct,
                        "数据来源": outcome.source,
                        "缓存命中": outcome.from_cache,
                    }
                    for daily_return in outcome.daily_returns
                )
            except Exception as exc:
                failure_rows.append(
                    {
                        "序号": company["序号"],
                        "股票代码": company["股票代码"],
                        "股票名称": company["股票名称"],
                        "失败原因": str(exc),
                    }
                )
            if completed == len(records) or completed % 10 == 0:
                print(
                    f"已处理 {completed}/{len(records)}；成功 {len(summary_rows)}；"
                    f"失败 {len(failure_rows)}；缓存命中 {cache_hits}",
                    flush=True,
                )

    summary_results = pd.DataFrame(summary_rows, columns=SUMMARY_COLUMNS)
    daily_results = pd.DataFrame(daily_rows, columns=DAILY_RETURN_COLUMNS)
    failures = pd.DataFrame(failure_rows, columns=FAILURE_COLUMNS)
    if not summary_results.empty:
        summary_results = summary_results.sort_values(
            ["区间涨跌幅（%）", "序号"],
            ascending=[False, True],
            kind="stable",
        ).reset_index(drop=True)
    if not daily_results.empty:
        daily_results = daily_results.sort_values(
            ["交易日期", "序号"],
            ascending=[True, True],
            kind="stable",
        ).reset_index(drop=True)
    if not failures.empty:
        failures = failures.sort_values("序号", kind="stable").reset_index(drop=True)
    return summary_results, daily_results, failures, {
        "总数": len(records),
        "成功": len(summary_results),
        "失败": len(failures),
        "缓存命中": cache_hits,
        "每日记录": len(daily_results),
    }


def build_wide_results(summary_results: pd.DataFrame, daily_results: pd.DataFrame) -> pd.DataFrame:
    """将逐日长表转换为一行一股、每列一个交易日的涨跌幅宽表。"""

    base = summary_results.reindex(
        columns=("序号", "股票代码", "股票名称", "数据来源", "缓存命中")
    ).copy()
    if base.empty:
        return pd.DataFrame(columns=WIDE_RETURN_FIXED_COLUMNS)

    base["股票代码"] = base["股票代码"].astype(str)
    base.insert(3, "总涨跌幅（%）", pd.NA)
    if daily_results.empty:
        return base.reindex(columns=WIDE_RETURN_FIXED_COLUMNS)

    daily = daily_results.reindex(
        columns=("股票代码", "交易日期", "当日涨跌幅（%）")
    ).copy()
    daily["股票代码"] = daily["股票代码"].astype(str)
    daily["交易日期"] = pd.to_datetime(daily["交易日期"], errors="coerce")
    daily = daily.dropna(subset=["交易日期"])
    if daily.empty:
        return base.reindex(columns=WIDE_RETURN_FIXED_COLUMNS)

    daily["_交易日列名"] = daily["交易日期"].dt.strftime("%Y-%m-%d")
    date_columns = sorted(daily["_交易日列名"].unique().tolist())

    def compounded_change_pct(values: pd.Series) -> float | None:
        numeric_values = pd.to_numeric(values, errors="coerce")
        if numeric_values.empty or numeric_values.isna().any():
            return None
        return round(float(((numeric_values / 100.0 + 1.0).prod() - 1.0) * 100.0), 10)

    totals = (
        daily.groupby("股票代码", sort=False)["当日涨跌幅（%）"]
        .agg(compounded_change_pct)
        .rename("总涨跌幅（%）")
        .reset_index()
    )
    pivoted = (
        daily.pivot(index="股票代码", columns="_交易日列名", values="当日涨跌幅（%）")
        .reindex(columns=date_columns)
        .reset_index()
    )
    wide_results = base.drop(columns="总涨跌幅（%）").merge(
        totals,
        on="股票代码",
        how="left",
        sort=False,
    )
    wide_results = wide_results.merge(pivoted, on="股票代码", how="left", sort=False)
    wide_results = wide_results.reindex(columns=(*WIDE_RETURN_FIXED_COLUMNS, *date_columns))
    return wide_results.sort_values(
        ["总涨跌幅（%）", "序号"],
        ascending=[False, True],
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)


def _is_daily_return_column(header: object) -> bool:
    """判断宽表表头是否为 ISO 格式的交易日。"""

    if not isinstance(header, str):
        return False
    try:
        date.fromisoformat(header)
    except ValueError:
        return False
    return True


def _format_output_sheet(worksheet: Any, *, freeze_panes: str = "A2") -> None:
    """设置报表工作表的筛选、冻结窗格、列宽和数值格式。"""

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = freeze_panes
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = worksheet.dimensions
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[column_letter].width = min(max(width, 12), 36)

    if "股票代码" in headers:
        for cell in worksheet[get_column_letter(headers["股票代码"])][1:]:
            cell.number_format = "@"
    for header in (
        "交易日期",
        "前一交易日",
        "请求开始日期",
        "实际开始交易日",
        "请求结束日期",
        "实际结束交易日",
    ):
        if header in headers:
            for cell in worksheet[get_column_letter(headers[header])][1:]:
                cell.number_format = "yyyy-mm-dd"
    for header in (
        "前收盘价（前复权）",
        "收盘价（前复权）",
        "开始收盘价（前复权）",
        "结束收盘价（前复权）",
    ):
        if header in headers:
            for cell in worksheet[get_column_letter(headers[header])][1:]:
                cell.number_format = "0.0000"
    percent_headers = {
        "当日涨跌幅（%）",
        "区间涨跌幅（%）",
        "总涨跌幅（%）",
    }
    for header, column in headers.items():
        if header in percent_headers or _is_daily_return_column(header):
            for cell in worksheet[get_column_letter(column)][1:]:
                cell.number_format = "0.00"


def _write_daily_detail_sheets(writer: Any, daily_results: pd.DataFrame) -> None:
    """写入每日长表；超过 Excel 行上限时自动拆分工作表。"""

    if len(daily_results) <= MAX_EXCEL_DATA_ROWS:
        daily_results.reindex(columns=DAILY_RETURN_COLUMNS).to_excel(
            writer,
            sheet_name="每日涨跌幅明细",
            index=False,
        )
        _format_output_sheet(writer.sheets["每日涨跌幅明细"])
        return

    for index, start_row in enumerate(range(0, len(daily_results), MAX_EXCEL_DATA_ROWS), start=1):
        sheet_name = f"每日涨跌幅明细_{index}"
        chunk = daily_results.iloc[start_row : start_row + MAX_EXCEL_DATA_ROWS]
        chunk.reindex(columns=DAILY_RETURN_COLUMNS).to_excel(writer, sheet_name=sheet_name, index=False)
        _format_output_sheet(writer.sheets[sheet_name])


def write_results_workbook(
    summary_results: pd.DataFrame,
    daily_results: pd.DataFrame,
    failures: pd.DataFrame,
    output_path: Path,
) -> None:
    """将每日宽表、每日明细、区间汇总和失败原因写入一个 Excel 文件。"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        wide_results = build_wide_results(summary_results, daily_results)
        wide_results.to_excel(writer, sheet_name="每日涨跌幅", index=False)
        _format_output_sheet(writer.sheets["每日涨跌幅"], freeze_panes="G2")
        _write_daily_detail_sheets(writer, daily_results)
        summary_results.reindex(columns=SUMMARY_COLUMNS).to_excel(
            writer,
            sheet_name="区间汇总",
            index=False,
        )
        failures.reindex(columns=FAILURE_COLUMNS).to_excel(
            writer,
            sheet_name="失败明细",
            index=False,
        )
        _format_output_sheet(writer.sheets["区间汇总"])
        _format_output_sheet(writer.sheets["失败明细"])


def default_output_path(requested_start: date, requested_end: date) -> Path:
    """生成默认报表文件名。"""

    return MODULE_DIR / f"深市主板每日涨跌幅_{requested_start.isoformat()}_{requested_end.isoformat()}.xlsx"


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """定义并解析命令行参数。"""

    parser = argparse.ArgumentParser(
        description="读取深交所主板股票池，导出指定期间每个交易日的前复权涨跌幅。"
    )
    parser.add_argument("--start-date", type=parse_iso_date, required=True, help="开始日期（YYYY-MM-DD）。")
    parser.add_argument("--end-date", type=parse_iso_date, required=True, help="结束日期（YYYY-MM-DD）。")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="股票池 Excel 路径（默认：深交所数据.xlsx）。",
    )
    parser.add_argument("--output", type=Path, help="输出 Excel 路径。")
    parser.add_argument(
        "--max-companies",
        type=int,
        help="仅处理前 N 只股票，用于快速测试。",
    )
    parser.add_argument(
        "--cache-hours",
        type=float,
        default=DEFAULT_CACHE_HOURS,
        help="完整日线缓存有效小时数（默认：12）。设为 0 可禁用读写缓存。",
    )
    parser.add_argument("--force-refresh", action="store_true", help="忽略已有缓存并重新获取行情。")
    parser.add_argument(
        "--workers",
        type=int,
        default=DEFAULT_WORKERS,
        help="并发请求数（默认：4，最大：6）。",
    )
    parser.add_argument(
        "--interval",
        type=float,
        default=DEFAULT_REQUEST_INTERVAL_SECONDS,
        help="所有请求共享的最小间隔秒数（默认：0.25）。",
    )
    parser.add_argument("--timeout", type=float, default=15.0, help="单次请求超时秒数（默认：15）。")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    """运行命令行任务并返回进程退出码。"""

    args = parse_args(argv)
    if args.start_date > args.end_date:
        raise ValueError("--start-date 不能晚于 --end-date。")
    if args.max_companies is not None and args.max_companies <= 0:
        raise ValueError("--max-companies 必须为正数。")
    if not 1 <= args.workers <= MAX_WORKERS:
        raise ValueError(f"--workers 必须在 1 到 {MAX_WORKERS} 之间。")
    if args.interval < 0:
        raise ValueError("--interval 不能为负数。")
    if args.timeout <= 0:
        raise ValueError("--timeout 必须为正数。")
    if args.cache_hours < 0:
        raise ValueError("--cache-hours 不能为负数。")

    companies = load_mainboard_companies(args.workbook)
    output_path = (args.output or default_output_path(args.start_date, args.end_date)).resolve()
    company_count = len(companies) if args.max_companies is None else min(len(companies), args.max_companies)
    print(
        f"正在获取 {company_count} 只股票在 {args.start_date} 至 {args.end_date} "
        "每个交易日的前复权涨跌幅..."
    )
    summary_results, daily_results, failures, summary = collect_period_returns(
        companies,
        requested_start=args.start_date,
        requested_end=args.end_date,
        max_companies=args.max_companies,
        cache_hours=args.cache_hours,
        force_refresh=args.force_refresh,
        workers=args.workers,
        request_interval_seconds=args.interval,
        timeout_seconds=args.timeout,
    )
    write_results_workbook(summary_results, daily_results, failures, output_path)
    print(f"已保存：{output_path}")
    print(
        f"统计：总数={summary['总数']}，成功={summary['成功']}，失败={summary['失败']}，"
        f"缓存命中={summary['缓存命中']}，每日记录={summary['每日记录']}"
    )
    return 0 if summary["成功"] else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, PeriodReturnError, ValueError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1) from exc
